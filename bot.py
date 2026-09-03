# ==========================================================================
#  AUTOMATIZADOR INVIMA
#  Soporte Multi-Proceso (Información General, Composición, etc.),
#  Búsqueda Exacta 1:1, 3 Reintentos por Fila, Continuidad de Lote y Reporte
# ==========================================================================

import sys
import os
import time
import re
import threading
import importlib.util
import urllib.request
import json
import subprocess
import hashlib
import uuid
import base64
import winreg
from datetime import datetime
import socket
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

PUERTO_CHROME = 9222   # Puerto de depuración de Chrome

VERSION_ACTUAL = "v1.1.6"
URL_VERSION_GITHUB = "https://raw.githubusercontent.com/Danielcastro5/bot-invima/main/version.json"
FIREBASE_DB_URL = "https://bot-invima-licencias-default-rtdb.firebaseio.com"
SECRET_SALT_LICENCIA = "BOT_INVIMA_SECURE_AUTH_SALT_2026_V1"

# Configuración inicial de CustomTkinter
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


# --------------------------------------------------------------------------
#  LANZADOR AUTOMÁTICO DE GOOGLE CHROME EN MODO DEPURACIÓN (PUERTO 9222)
# --------------------------------------------------------------------------
def esta_puerto_abierto(puerto=9222):
    try:
        with socket.create_connection(("127.0.0.1", puerto), timeout=1):
            return True
    except Exception:
        return False


def abrir_chrome_automatizado(app=None):
    """
    Inicia Google Chrome en modo de depuración remota (puerto 9222) de forma transparente.
    """
    if esta_puerto_abierto(PUERTO_CHROME):
        if app:
            app.log("✅ Chrome automatizado ya se encuentra en ejecución en puerto 9222.", "info")
        return True

    rutas_chrome = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.join(os.environ.get("LOCALAPPDATA", ""), r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES", ""), r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", ""), r"Google\Chrome\Application\chrome.exe")
    ]

    exe_chrome = None
    for r in rutas_chrome:
        if os.path.exists(r):
            exe_chrome = r
            break

    if not exe_chrome:
        if app:
            app.log("❌ No se encontró Google Chrome en las rutas predeterminadas.", "error")
        return False

    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    user_data_dir = os.path.join(base_dir, "chrome_profile_bot")

    cmd = [
        exe_chrome,
        f"--remote-debugging-port={PUERTO_CHROME}",
        f"--user-data-dir={user_data_dir}"
    ]

    try:
        subprocess.Popen(cmd)
        time.sleep(2)
        if app:
            app.log("🚀 Chrome automatizado iniciado correctamente (Puerto 9222).", "success")
            app.log("💡 Inicia sesión en el portal INVIMA dentro de esa ventana de Chrome.", "warning")
        return True
    except Exception as e:
        if app:
            app.log(f"❌ Error al lanzar Chrome automáticamente: {e}", "error")
        return False



# --------------------------------------------------------------------------
#  SISTEMA DE LICENCIAMIENTO SEGURO CON HWID INMUTABLE & FIREBASE
# --------------------------------------------------------------------------
def obtener_hwid():
    """
    Genera un identificador 100% permanente e inalterable del computador (HWID).
    Utiliza el MachineGuid del registro de Windows y el UUID del sistema (Motherboard/BIOS).
    No depende de la red, Wi-Fi, Ethernet, VPN ni conexiones de internet.
    """
    identificadores = []

    # 1. MachineGuid del Registro de Windows (Inmutable por instalación)
    try:
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography", 0, winreg.KEY_READ | winreg.KEY_WOW64_64KEY) as key:
            machine_guid, _ = winreg.QueryValueEx(key, "MachineGuid")
            if machine_guid and str(machine_guid).strip():
                identificadores.append(str(machine_guid).strip())
    except Exception:
        try:
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography", 0, winreg.KEY_READ) as key:
                machine_guid, _ = winreg.QueryValueEx(key, "MachineGuid")
                if machine_guid and str(machine_guid).strip():
                    identificadores.append(str(machine_guid).strip())
        except Exception:
            pass

    # 2. UUID del Hardware / Placa Base (PowerShell / CIM / WMIC)
    try:
        cmd = 'powershell -NoProfile -Command "(Get-CimInstance Win32_ComputerSystemProduct).UUID"'
        out = subprocess.check_output(cmd, shell=True, timeout=3).decode().strip()
        if out and len(out) > 8 and "error" not in out.lower():
            identificadores.append(out)
    except Exception:
        try:
            cmd = "wmic csproduct get uuid"
            out = subprocess.check_output(cmd, shell=True, timeout=3).decode().split('\n')
            if len(out) > 1 and out[1].strip():
                identificadores.append(out[1].strip())
        except Exception:
            pass

    # 3. Nombre del Computador
    nombre_pc = os.getenv("COMPUTERNAME", "PC_DEFAULT")
    identificadores.append(nombre_pc)

    cadena_unica = "_".join(identificadores)
    return hashlib.sha256(cadena_unica.encode("utf-8")).hexdigest()[:16].upper()


def _generar_clave_cifrado():
    hwid = obtener_hwid()
    key_material = f"{hwid}_{SECRET_SALT_LICENCIA}"
    return hashlib.sha256(key_material.encode("utf-8")).digest()


def _cifrar_datos_licencia(texto_str):
    try:
        key = _generar_clave_cifrado()
        raw_bytes = texto_str.encode("utf-8")
        encrypted = bytearray()
        for i, b in enumerate(raw_bytes):
            k = key[i % len(key)]
            encrypted.append(b ^ k)
        hmac_sig = hashlib.sha256(key + bytes(encrypted)).hexdigest()[:16]
        payload = json.dumps({"sig": hmac_sig, "data": base64.b64encode(bytes(encrypted)).decode()})
        return base64.b64encode(payload.encode()).decode()
    except Exception:
        return ""


def _descifrar_datos_licencia(payload_b64):
    try:
        key = _generar_clave_cifrado()
        raw_json = base64.b64decode(payload_b64.encode()).decode()
        payload = json.loads(raw_json)
        enc_bytes = base64.b64decode(payload["data"].encode())
        expected_sig = hashlib.sha256(key + enc_bytes).hexdigest()[:16]
        if payload.get("sig") != expected_sig:
            return None  # Alterado o pertenece a otro equipo!
        decrypted = bytearray()
        for i, b in enumerate(enc_bytes):
            k = key[i % len(key)]
            decrypted.append(b ^ k)
        return decrypted.decode("utf-8")
    except Exception:
        return None


def obtener_ruta_licencia_local():
    """Retorna la ruta del archivo local cifrado donde se guarda la licencia activada."""
    base = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
    carpeta = os.path.join(base, "BotINVIMA_Data")
    os.makedirs(carpeta, exist_ok=True)
    return os.path.join(carpeta, "license.dat")


def guardar_licencia_local(clave, info_dict):
    try:
        ruta = obtener_ruta_licencia_local()
        datos = {
            "clave": clave,
            "empresa": info_dict.get("empresa", ""),
            "hwid": obtener_hwid(),
            "vencimiento": info_dict.get("vencimiento", "")
        }
        contenido_cifrado = _cifrar_datos_licencia(json.dumps(datos))
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(contenido_cifrado)

        # Eliminar archivo legacy en texto plano si existía
        ruta_legacy = os.path.join(os.path.dirname(ruta), "license.json")
        if os.path.exists(ruta_legacy):
            try:
                os.remove(ruta_legacy)
            except Exception:
                pass
    except Exception as e:
        print(f"Error guardando licencia local cifrada: {e}")


def leer_licencia_local():
    try:
        ruta = obtener_ruta_licencia_local()
        if os.path.exists(ruta):
            with open(ruta, "r", encoding="utf-8") as f:
                cifrado = f.read().strip()
            plano = _descifrar_datos_licencia(cifrado)
            if plano:
                return json.loads(plano)
    except Exception:
        pass
    return None


def eliminar_licencia_local():
    """Elimina los archivos de licencia guardados localmente."""
    try:
        ruta = obtener_ruta_licencia_local()
        if os.path.exists(ruta):
            os.remove(ruta)
        ruta_legacy = os.path.join(os.path.dirname(ruta), "license.json")
        if os.path.exists(ruta_legacy):
            try:
                os.remove(ruta_legacy)
            except Exception:
                pass
    except Exception:
        pass


def validar_licencia_firebase(clave_licencia):
    """
    Verifica la clave de licencia en Firebase Realtime Database de forma segura.
    Soporta desvinculación remota, anti-duplicación automática y bloqueo por HWID permanente.
    """
    clave = clave_licencia.strip().upper()
    if not clave:
        return False, "Por favor ingresa una clave de licencia."

    hwid = obtener_hwid()
    nombre_pc_actual = os.getenv("COMPUTERNAME", "PC_ACTUAL").strip().upper()
    url = f"{FIREBASE_DB_URL}/licencias/{clave}.json"

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw_data = resp.read().decode("utf-8")
            if not raw_data or raw_data == "null":
                eliminar_licencia_local()
                return False, f"La clave de licencia '{clave}' no existe o fue eliminada."

            data = json.loads(raw_data)

        if not data.get("activa", False):
            eliminar_licencia_local()
            return False, "Licencia inactiva o suspendida por el proveedor."

        vencimiento = data.get("vencimiento", "")
        if vencimiento:
            try:
                fecha_venc = datetime.strptime(vencimiento, "%Y-%m-%d")
                if datetime.now() > fecha_venc:
                    eliminar_licencia_local()
                    return False, f"Licencia vencida el {vencimiento}. Contacta al soporte para renovar."
            except Exception:
                pass

        equipos = data.get("equipos", {})
        if not isinstance(equipos, dict):
            equipos = {}

        max_equipos = int(data.get("max_equipos", 1))

        # 1. Si el HWID inmutable ya está registrado en Firebase
        if hwid in equipos:
            val_equipo = equipos[hwid]
            if val_equipo is False or str(val_equipo).lower() == "false":
                eliminar_licencia_local()
                return False, "Este equipo específico ha sido deshabilitado por el administrador."

            # Actualizar nombre si fuera necesario
            if val_equipo != nombre_pc_actual and val_equipo is not True:
                try:
                    url_put = f"{FIREBASE_DB_URL}/licencias/{clave}/equipos/{hwid}.json"
                    req_put = urllib.request.Request(url_put, data=json.dumps(nombre_pc_actual).encode("utf-8"), headers={"Content-Type": "application/json"}, method="PUT")
                    with urllib.request.urlopen(req_put, timeout=5): pass
                except Exception: pass

            equipos_activos = [k for k, v in equipos.items() if v is not False and str(v).lower() != "false"]
            info = {
                "clave": clave,
                "empresa": data.get("empresa", "Cliente"),
                "vencimiento": vencimiento or "Permanente",
                "equipos_usados": len(equipos_activos),
                "max_equipos": max_equipos
            }
            guardar_licencia_local(clave, info)
            return True, info

        # 2. Anti-duplicación inteligente:
        #    Si este mismo nombre de computador ya existía bajo una clave vieja generada por el adaptador anterior,
        #    migramos esa clave al HWID inmutable y eliminamos el duplicado anterior sin consumir un nuevo cupo.
        hwid_antiguo_encontrado = None
        for k_old, v_name in equipos.items():
            if str(v_name).strip().upper() == nombre_pc_actual and v_name is not False and str(v_name).lower() != "false":
                hwid_antiguo_encontrado = k_old
                break

        if hwid_antiguo_encontrado:
            # Eliminar el HWID antiguo de Firebase
            try:
                url_del = f"{FIREBASE_DB_URL}/licencias/{clave}/equipos/{hwid_antiguo_encontrado}.json"
                req_del = urllib.request.Request(url_del, headers={"User-Agent": "Mozilla/5.0"}, method="DELETE")
                with urllib.request.urlopen(req_del, timeout=5): pass
            except Exception: pass

            # Registrar el nuevo HWID inmutable
            try:
                url_put = f"{FIREBASE_DB_URL}/licencias/{clave}/equipos/{hwid}.json"
                req_put = urllib.request.Request(url_put, data=json.dumps(nombre_pc_actual).encode("utf-8"), headers={"Content-Type": "application/json"}, method="PUT")
                with urllib.request.urlopen(req_put, timeout=5): pass
            except Exception: pass

            equipos[hwid] = nombre_pc_actual
            if hwid_antiguo_encontrado in equipos:
                del equipos[hwid_antiguo_encontrado]

            equipos_activos = [k for k, v in equipos.items() if v is not False and str(v).lower() != "false"]
            info = {
                "clave": clave,
                "empresa": data.get("empresa", "Cliente"),
                "vencimiento": vencimiento or "Permanente",
                "equipos_usados": len(equipos_activos),
                "max_equipos": max_equipos
            }
            guardar_licencia_local(clave, info)
            return True, info

        # 3. Si el HWID NO está en Firebase, pero el PC conservaba una licencia local previa:
        datos_locales = leer_licencia_local()
        if datos_locales and datos_locales.get("clave") == clave:
            eliminar_licencia_local()
            return False, "Este equipo ha sido desvinculado de la licencia por el administrador."

        # 4. Registro de un computador nuevo en la licencia
        equipos_activos = [k for k, v in equipos.items() if v is not False and str(v).lower() != "false"]
        if len(equipos_activos) >= max_equipos:
            return False, f"Límite de dispositivos alcanzado (Máximo {max_equipos} equipo(s) para esta licencia)."

        # Registrar este nuevo equipo (HWID) en Firebase
        url_put = f"{FIREBASE_DB_URL}/licencias/{clave}/equipos/{hwid}.json"
        body = json.dumps(nombre_pc_actual).encode("utf-8")
        req_put = urllib.request.Request(url_put, data=body, headers={"Content-Type": "application/json"}, method="PUT")
        with urllib.request.urlopen(req_put, timeout=10) as resp_put:
            pass

        info = {
            "clave": clave,
            "empresa": data.get("empresa", "Cliente"),
            "vencimiento": vencimiento or "Permanente",
            "equipos_usados": len(equipos_activos) + 1,
            "max_equipos": max_equipos
        }
        guardar_licencia_local(clave, info)
        return True, info

    except Exception as e:
        return False, f"Error al verificar la licencia en línea: {e}"



# --------------------------------------------------------------------------
#  SISTEMA DE AUTO-ACTUALIZACIÓN DESDE GITHUB (REMOTA Y SILENCIOSA)
# --------------------------------------------------------------------------
def buscar_actualizaciones_github(app):
    """
    Consulta en segundo plano si existe una versión más reciente en GitHub.
    """
    def _tarea():
        try:
            time.sleep(2)  # Esperar que la ventana cargue completamente
            req = urllib.request.Request(
                URL_VERSION_GITHUB,
                headers={"User-Agent": "Mozilla/5.0"}
            )
            with urllib.request.urlopen(req, timeout=6) as resp:
                data = json.loads(resp.read().decode("utf-8"))

            version_remota = str(data.get("version", "")).strip()
            if not version_remota:
                return

            if version_remota != VERSION_ACTUAL and not version_remota.startswith(VERSION_ACTUAL):
                app.log(f"🔔 ¡NUEVA VERSIÓN DISPONIBLE EN GITHUB! ({version_remota})", "warning")
                app.mostrar_modal_actualizacion(data)
            else:
                app.log(f"✅ Bot actualizado a la versión oficial ({VERSION_ACTUAL}).", "info")
        except Exception:
            pass

    t = threading.Thread(target=_tarea, daemon=True)
    t.start()


def ejecutar_actualizacion_automatica(exe_url, config_url, app):
    """
    Descarga la nueva versión desde GitHub y reinicia la aplicación automáticamente.
    """
    try:
        app.log("⬇️ Descargando actualización desde GitHub...", "info")
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        ruta_exe_nuevo = os.path.join(base_dir, "Automatizador_INVIMA_nueva.exe")
        ruta_exe_actual = os.path.join(base_dir, "Automatizador INVIMA.exe")
        ruta_config = os.path.join(base_dir, "config.py")

        # 1. Descargar nuevo ejecutable
        if exe_url:
            req_exe = urllib.request.Request(exe_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req_exe, timeout=60) as resp, open(ruta_exe_nuevo, "wb") as out_file:
                out_file.write(resp.read())

        if os.path.exists(ruta_exe_nuevo) and os.path.getsize(ruta_exe_nuevo) < 10 * 1024 * 1024:
            if os.path.exists(ruta_exe_nuevo):
                os.remove(ruta_exe_nuevo)
            raise Exception("El archivo de actualización descargado está incompleto o dañado.")

        # 2. Descargar nuevo config.py
        if config_url:
            try:
                req_cfg = urllib.request.Request(config_url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req_cfg, timeout=15) as resp, open(ruta_config, "wb") as out_cfg:
                    out_cfg.write(resp.read())
            except Exception:
                pass

        # 3. Crear script de reemplazo en segundo plano
        ruta_bat = os.path.join(base_dir, "actualizar_bot.bat")
        script_bat = f"""@echo off
timeout /t 3 /nobreak > nul
taskkill /F /IM "Automatizador INVIMA.exe" 2>nul
timeout /t 1 /nobreak > nul
if exist "{ruta_exe_nuevo}" (
    move /y "{ruta_exe_nuevo}" "{ruta_exe_actual}"
    start "" "{ruta_exe_actual}"
)
del "%~f0"
"""
        with open(ruta_bat, "w", encoding="utf-8") as f:
            f.write(script_bat)

        app.log("✨ Descarga completada. Reiniciando bot...", "success")
        time.sleep(1)

        # 4. Iniciar script y cerrar aplicación actual
        subprocess.Popen(["cmd.exe", "/c", ruta_bat], creationflags=subprocess.CREATE_NO_WINDOW)
        os._exit(0)

    except Exception as e:
        app.log(f"❌ Error al aplicar actualización automática: {e}", "error")
        messagebox.showerror("Error de Actualización", f"No se pudo descargar la actualización:\n{e}")


# --------------------------------------------------------------------------
#  Carga Dinámica de config.py
# --------------------------------------------------------------------------
def cargar_config_dinamico():
    """Carga siempre la versión actual de config.py guardada en la carpeta del ejecutable/script."""
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    ruta_config = os.path.join(base_dir, "config.py")

    if not os.path.exists(ruta_config):
        ruta_config = os.path.join(os.getcwd(), "config.py")

    if os.path.exists(ruta_config):
        try:
            spec = importlib.util.spec_from_file_location("config_user", ruta_config)
            cfg = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg)
            return cfg
        except Exception as e:
            print(f"Error cargando config.py local: {e}")

    import config as cfg_fallback
    return cfg_fallback


def obtener_dict_proceso(cfg, nombre_proceso):
    """Retorna el diccionario de configuración del proceso seleccionado."""
    if hasattr(cfg, "PROCESOS") and isinstance(cfg.PROCESOS, dict) and nombre_proceso in cfg.PROCESOS:
        return cfg.PROCESOS[nombre_proceso]
    
    # Fallback si config.py no usa PROCESOS
    return {
        "NOMBRE_HOJA": getattr(cfg, "NOMBRE_HOJA", "Matriz Presentaciones"),
        "BOTON_ABRIR_MODAL": getattr(cfg, "BOTON_ABRIR_MODAL", ""),
        "SELECTOR_MODAL": getattr(cfg, "SELECTOR_MODAL", ""),
        "BOTON_ENVIAR": getattr(cfg, "BOTON_ENVIAR", ""),
        "CAMPOS": getattr(cfg, "CAMPOS", [])
    }


# --------------------------------------------------------------------------
#  Normalización de cadenas (ignora tildes, acentos, mayúsculas y espacios extra)
# --------------------------------------------------------------------------
import unicodedata

def normalizar_texto(texto):
    if not texto:
        return ""
    s = unicodedata.normalize('NFD', str(texto).strip().lower())
    return "".join(c for c in s if unicodedata.category(c) != 'Mn')


# --------------------------------------------------------------------------
#  Lectura y validación del Excel para el proceso activo
# --------------------------------------------------------------------------
def leer_excel(ruta, app, cfg, proceso_config):
    try:
        libro = load_workbook(ruta, data_only=True)
    except FileNotFoundError:
        app.log("❌ ERROR: No se encontró el archivo de Excel.", "error")
        return None
    except Exception as e:
        app.log(f"❌ ERROR al abrir Excel: {e}", "error")
        return None

    nombre_hoja = proceso_config.get("NOMBRE_HOJA", "Matriz Presentaciones")
    
    # Búsqueda flexible de la hoja (exacta o ignorando tildes y mayúsculas)
    hoja_encontrada = None
    if nombre_hoja in libro.sheetnames:
        hoja_encontrada = nombre_hoja
    else:
        norm_hoja_req = normalizar_texto(nombre_hoja)
        for s in libro.sheetnames:
            if normalizar_texto(s) == norm_hoja_req:
                hoja_encontrada = s
                break

    if not hoja_encontrada:
        app.log(f"❌ ERROR: El Excel seleccionado no contiene la hoja '{nombre_hoja}'.", "error")
        app.log(f"📋 Hojas disponibles en este Excel: {', '.join(libro.sheetnames)}", "warning")
        app.log(f"💡 Asegúrate de que el Excel contenga una pestaña llamada '{nombre_hoja}'", "info")
        return None

    hoja = libro[hoja_encontrada]
    crudas = list(hoja.iter_rows(values_only=True))
    if len(crudas) < 2:
        app.log(f"❌ ERROR: La hoja '{hoja_encontrada}' no contiene registros suficientes.", "error")
        return None

    encabezados = [str(c).strip() if c is not None else "" for c in crudas[0]]
    campos_req = proceso_config.get("CAMPOS_CABECERA", []) + proceso_config.get("CAMPOS", [])
    
    # Mapeo de columnas requeridas con soporte de alias y normalización
    mapa_columnas = {}
    faltan = []

    for c in campos_req:
        col_nombre = c["columna"]
        alias_lista = [col_nombre] + c.get("alias", [])
        encontrado_idx = None

        for idx, enc in enumerate(encabezados):
            norm_enc = normalizar_texto(enc)
            for a in alias_lista:
                if enc == a or norm_enc == normalizar_texto(a):
                    encontrado_idx = idx
                    break
            if encontrado_idx is not None:
                break

        if encontrado_idx is not None:
            mapa_columnas[col_nombre] = encontrado_idx
        else:
            faltan.append(col_nombre)

    col_extra = proceso_config.get("COLUMNA_NOMBRE_FORMULA")
    if col_extra:
        encontrado_extra = None
        for idx, enc in enumerate(encabezados):
            if enc == col_extra or normalizar_texto(enc) == normalizar_texto(col_extra) or "formula" in normalizar_texto(enc):
                encontrado_extra = idx
                break
        if encontrado_extra is not None:
            mapa_columnas[col_extra] = encontrado_extra
        elif col_extra not in faltan:
            faltan.append(col_extra)

    if faltan:
        app.log(f"❌ ERROR: Faltan las siguientes columnas en la hoja '{hoja_encontrada}':", "error")
        for c in faltan:
            app.log(f"   • {c}", "error")
        return None

    filas = []
    for i, cruda in enumerate(crudas[1:], start=2): # i representa la línea real en Excel
        if all(v is None for v in cruda):
            continue
        fila = {"__linea_excel__": i}
        for col_nombre, col_idx in mapa_columnas.items():
            v = cruda[col_idx] if col_idx < len(cruda) else None
            fila[col_nombre] = "" if v is None else str(v).strip()
        filas.append(fila)

    app.log(f"✅ Excel validado correctamente para la hoja '{hoja_encontrada}': {len(filas)} filas cargadas.", "success")
    return filas


def hacer_clic_opcion(opcion_locator):
    try:
        opcion_locator.scroll_into_view_if_needed(timeout=1000)
    except Exception:
        pass
    try:
        hijo_txt = opcion_locator.locator(".ant-select-item-option-content")
        if hijo_txt.count() > 0 and hijo_txt.first.is_visible():
            hijo_txt.first.click(force=True, timeout=1500)
            return
    except Exception:
        pass
    try:
        opcion_locator.click(force=True, timeout=1500)
    except Exception:
        try:
            opcion_locator.evaluate("el => el.click()")
        except Exception:
            pass


def obtener_elemento_visible(page, selector_str):
    """Escanea todos los elementos coincidentes y devuelve el primero visible."""
    locs = page.locator(selector_str)
    count = locs.count()
    if count == 0:
        return None
    for i in range(count):
        el = locs.nth(i)
        try:
            if el.is_visible():
                return el
        except Exception:
            pass
    return locs.first


# --------------------------------------------------------------------------
#  MANEJADOR EXCLUSIVO Y RIGUROSO PARA INFORMACIÓN GENERAL (PRESENTACIONES)
# --------------------------------------------------------------------------

def _limpiar_e_ingresar_texto(page, target, valor):
    """
    Enfoca la casilla específica y tipea el texto de forma segura sin tocar el teclado global.
    """
    try:
        target.click(force=True, timeout=1000)
    except Exception:
        try:
            target.focus(timeout=1000)
        except Exception:
            pass

    time.sleep(0.05)

    try:
        target.fill("")
    except Exception:
        try:
            target.press("Control+a")
            target.press("Backspace")
        except Exception:
            pass

    time.sleep(0.05)

    try:
        target.press_sequentially(str(valor), delay=40)
    except Exception:
        try:
            target.fill(str(valor))
        except Exception:
            pass

    time.sleep(0.25)


def _seleccionar_opcion_antdesign(page, target, valor, app):
    """
    Despliega y selecciona la opción exacta en Ant Design Select.
    Prioriza opciones de texto real sobre códigos internos.
    """
    valor_norm = normalizar_texto(valor)
    valor_str = str(valor).strip()

    # 1. Esperar menú desplegable
    try:
        page.wait_for_selector(".ant-select-dropdown:not(.ant-select-dropdown-hidden)", timeout=1500)
    except Exception:
        pass

    dropdown = page.locator(".ant-select-dropdown:not(.ant-select-dropdown-hidden)").last
    opciones = dropdown.locator(".ant-select-item-option:not(.ant-select-item-option-disabled)")
    total = opciones.count()

    if total == 0:
        opciones = page.locator(".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option")
        total = opciones.count()

    # Log de opciones encontradas para transparencia total
    lista_textos = []
    for k in range(total):
        try:
            t = opciones.nth(k).inner_text().strip()
            if t and t not in lista_textos:
                lista_textos.append(t)
        except Exception:
            pass

    if lista_textos:
        app.log(f"      📋 Opciones en menú ({total}): [{', '.join(lista_textos[:6])}]", "detail")
    else:
        app.log(f"      ℹ️ No se desplegaron opciones en menú para '{valor}'", "detail")

    # Si hay opciones desplegadas
    if total > 0:
        # A) Coincidencia EXACTA DIRECTA (Case-sensitive: 'Envase' == 'Envase', evitando 'envase' y 'envas1')
        for k in range(total):
            try:
                opc = opciones.nth(k)
                txt = opc.inner_text().strip()
                if txt == valor_str:
                    app.log(f"      🎯 Clic en coincidencia EXACTA (Directa): '{txt}'", "detail")
                    hacer_clic_opcion(opc)
                    time.sleep(0.3)
                    return True
            except Exception:
                pass

        # B) Coincidencia EXACTA NORMALIZADA (ej: 'Cartón' vs 'carton', descartando sufijos como carto1)
        for k in range(total):
            try:
                opc = opciones.nth(k)
                txt = opc.inner_text().strip()
                if normalizar_texto(txt) == valor_norm and not (len(txt) > 3 and txt[-1].isdigit()):
                    app.log(f"      🎯 Clic en coincidencia EXACTA (Normalizada): '{txt}'", "detail")
                    hacer_clic_opcion(opc)
                    time.sleep(0.3)
                    return True
            except Exception:
                pass

        # C) Coincidencia PARCIAL (solo opciones reales sin sufijos de código como envas1)
        for k in range(total):
            try:
                opc = opciones.nth(k)
                txt = opc.inner_text().strip()
                txt_n = normalizar_texto(txt)
                if (txt_n.startswith(valor_norm) or valor_norm in txt_n) and not (len(txt) > 3 and txt[-1].isdigit()):
                    app.log(f"      🎯 Clic en coincidencia PARCIAL: '{txt}'", "detail")
                    hacer_clic_opcion(opc)
                    time.sleep(0.3)
                    return True
            except Exception:
                pass

    # D) Fallback con teclado ArrowDown + Enter
    try:
        target.press("ArrowDown")
        time.sleep(0.1)
        target.press("Enter")
        time.sleep(0.25)
        return True
    except Exception:
        pass

    return False


def _verificar_campo_seleccionado(page, target, campo, valor, app):
    """
    VERIFICACIÓN ESTRICTA DEL CAMPO:
    Para autocompletar, exige que .ant-select-selection-item esté visible y activo.
    """
    valor_norm = normalizar_texto(valor)
    col_nombre = campo.get("columna", "")
    tipo_campo = campo.get("tipo", "autocompletar")

    # 1. Buscar etiqueta .ant-select-selection-item dentro del .ant-select MAS CERCANO a este input
    try:
        contenedor_select = target.locator("xpath=ancestor::div[contains(@class,'ant-select')][1]")
        if contenedor_select.count() > 0:
            selection_item = contenedor_select.locator(".ant-select-selection-item").first
            if selection_item.count() > 0 and selection_item.is_visible():
                texto_item = selection_item.inner_text().strip()
                if texto_item:
                    app.log(f"      ✅ Selección confirmada en portal (.ant-select-selection-item): '{texto_item}'", "detail")
                    return True
    except Exception:
        pass

    # 2. Si es campo de tipo 'texto', verificar target.input_value()
    if tipo_campo == "texto":
        try:
            val = target.input_value()
            if val and str(val).strip():
                app.log(f"      ✅ Valor verificado en casilla '{col_nombre}': '{val.strip()}'", "detail")
                return True
        except Exception:
            pass

    return False


def llenar_campo_presentaciones(page, campo, fila, app, cfg):
    valor = fila.get(campo["columna"], "")
    if not valor or str(valor).strip() == "":
        return

    target = obtener_elemento_visible(page, campo["selector"])
    if not target:
        app.log(f"   ⚠️ No se encontró la casilla para '{campo['columna']}'", "warning")
        return

    app.log(f"   ➔ {campo['columna']}: '{valor}'", "detail")

    if campo["tipo"] == "texto":
        try:
            target.fill(str(valor))
            time.sleep(0.15)
        except Exception:
            _limpiar_e_ingresar_texto(page, target, valor)
        return

    # Para autocompletar: reintentos estrictos por campo sin cerrar la ventana emergente
    for intento in range(1, 4):
        _limpiar_e_ingresar_texto(page, target, valor)
        _seleccionar_opcion_antdesign(page, target, valor, app)

        time.sleep(0.3)

        if _verificar_campo_seleccionado(page, target, campo, valor, app):
            return  # ¡Éxito real confirmado!

        app.log(f"      ⚠️ Intento {intento}/3: Campo '{campo['columna']}' no confirmó el valor '{valor}'", "warning")
        # Desenfocar suavemente haciendo clic en la cabecera del modal para cerrar menús colgados sin cerrar el modal
        try:
            page.locator(".ant-modal-header, .ant-modal-title").first.click(force=True, timeout=500)
            time.sleep(0.15)
        except Exception:
            pass

    # Si tras 3 intentos no se confirmó la selección:
    raise ValueError(f"No se pudo seleccionar '{valor}' en '{campo['columna']}'. El portal no registró el valor.")


# --------------------------------------------------------------------------
#  MANEJADOR EXCLUSIVO PARA COMPOSICIÓN (INGREDIENTES Y MEZCLAS)
# --------------------------------------------------------------------------
def llenar_autocompletar_composicion(page, campo, valor, timeout_ms, app):
    selector = campo["selector"]
    valor_norm = normalizar_texto(valor)

    # Identificar el modal activo superior si existe
    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    if modal_activo.count() > 0:
        modal_top = modal_activo.last
        target = modal_top.locator(selector).last
        if target.count() == 0:
            target = modal_top.locator(f".ant-form-item:has-text('{campo['columna']}') input").last
        if target.count() == 0:
            target = page.locator(selector).last
    else:
        target = page.locator(selector).last
        if target.count() == 0:
            target = page.locator(selector).first

    sugerencia_sel = campo.get("selector_sugerencia") or ".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option"
    col_name = campo["columna"].lower()
    es_ingrediente = "ingrediente" in col_name or "mezcla" in col_name

    timeout_espera_servidor = 8000 if es_ingrediente else 4000
    delay_tipeo = 80
    pausa_post_clic = 0.8 if es_ingrediente else 0.5
    max_intentos_tipeo = 3 if es_ingrediente else 2

    menu_desplegado = False

    for intento_t in range(1, max_intentos_tipeo + 1):
        try:
            try:
                target.scroll_into_view_if_needed(timeout=1000)
            except Exception:
                pass
            try:
                target.click(force=True, timeout=2000)
                time.sleep(0.1)
                target.fill("")
                time.sleep(0.1)
            except Exception:
                pass

            if intento_t > 1:
                app.log(f"      🔄 Re-escribiendo '{campo['columna']}' (Intento {intento_t}/{max_intentos_tipeo})...", "warning")

            target.type(str(valor), delay=delay_tipeo)
            time.sleep(0.4)

            page.wait_for_selector(sugerencia_sel, timeout=timeout_espera_servidor)
            menu_desplegado = True
            break

        except PWTimeout:
            if es_ingrediente:
                app.log(f"      ⏳ Esperando respuesta del servidor de ingredientes para '{campo['columna']}'...", "detail")

    if not menu_desplegado:
        page.wait_for_selector(sugerencia_sel, timeout=timeout_ms)

    dropdown_activo = page.locator(".ant-select-dropdown:not(.ant-select-dropdown-hidden)").last
    opciones = dropdown_activo.locator(".ant-select-item-option, div[role='option'], .ant-select-item-option-content")
    total_opciones = opciones.count()

    if total_opciones == 0:
        opciones = page.locator(sugerencia_sel)
        total_opciones = opciones.count()

    if total_opciones == 0:
        raise ValueError(f"La lista desplegable de ingredientes no mostró opciones para '{valor}'")

    for k in range(total_opciones):
        txt_opcion = opciones.nth(k).inner_text()
        if normalizar_texto(txt_opcion) == valor_norm:
            app.log(f"      🎯 Encontrada coincidencia EXACTA para '{valor}' -> '{txt_opcion.strip()}'", "detail")
            hacer_clic_opcion(opciones.nth(k))
            time.sleep(pausa_post_clic)
            return

    for k in range(total_opciones):
        txt_opcion = opciones.nth(k).inner_text()
        txt_norm = normalizar_texto(txt_opcion)
        if txt_norm.startswith(valor_norm) or valor_norm in txt_norm:
            app.log(f"      🎯 Encontrada coincidencia semejante para '{valor}' -> '{txt_opcion.strip()}'", "detail")
            hacer_clic_opcion(opciones.nth(k))
            time.sleep(pausa_post_clic)
            return

    coincidencia = opciones.filter(has_text=valor)
    if coincidencia.count() > 0:
        hacer_clic_opcion(coincidencia.first)
        time.sleep(pausa_post_clic)
    elif total_opciones > 0:
        app.log(f"      🎯 Seleccionando opción en lista para '{valor}'", "detail")
        hacer_clic_opcion(opciones.first)
        time.sleep(pausa_post_clic)


def llenar_multiselect(page, campo, valor, timeout_ms, app):
    import re
    items = [x.strip() for x in re.split(r'[,;|\n]', str(valor)) if x.strip()]
    if not items:
        return

    selector = campo["selector"]
    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    if modal_activo.count() > 0:
        modal_top = modal_activo.last
        target = modal_top.locator(selector).last
        if target.count() == 0:
            target = modal_top.locator(f".ant-form-item:has-text('{campo['columna']}') input").last
        if target.count() == 0:
            target = modal_top.locator(f".ant-form-item:has-text('{campo['columna']}') .ant-select-selector").last
        if target.count() == 0:
            target = page.locator(selector).last
    else:
        modal_top = page
        target = page.locator(selector).last

    sugerencia_sel = campo.get("selector_sugerencia") or ".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option"

    for idx, item in enumerate(items):
        item_norm = normalizar_texto(item)
        app.log(f"      ➔ Seleccionando función ({idx+1}/{len(items)}): '{item}'", "detail")

        # Asegurar foco haciendo clic en el selector visual de la fila
        try:
            modal_top.locator(f".ant-form-item:has-text('{campo['columna']}') .ant-select-selector").last.click(force=True, timeout=1500)
            time.sleep(0.3)
        except Exception:
            try:
                target.click(force=True, timeout=1500)
                time.sleep(0.3)
            except Exception:
                pass

        for intento_t in range(1, 3):
            try:
                try:
                    target.fill("")
                    time.sleep(0.1)
                except Exception:
                    pass

                target.type(str(item), delay=60)
                time.sleep(0.4)

                timeout_intento = min(3500, timeout_ms)
                page.wait_for_selector(sugerencia_sel, timeout=timeout_intento)
                break
            except PWTimeout:
                try:
                    target.fill("")
                    time.sleep(0.2)
                    target.type(item, delay=80)
                except Exception:
                    pass

        dropdown_activo = page.locator(".ant-select-dropdown:not(.ant-select-dropdown-hidden)").last
        opciones = dropdown_activo.locator(".ant-select-item-option, div[role='option'], .ant-select-item-option-content")
        total = opciones.count()
        if total == 0:
            opciones = page.locator(sugerencia_sel)
            total = opciones.count()

        if total == 0:
            app.log(f"      ⚠️ No se hallaron opciones para '{item}'", "warning")
            continue

        encontrado = False
        for k in range(total):
            txt_opcion = opciones.nth(k).inner_text()
            if normalizar_texto(txt_opcion) == item_norm:
                hacer_clic_opcion(opciones.nth(k))
                encontrado = True
                time.sleep(0.4)
                break

        if not encontrado:
            for k in range(total):
                txt_opcion = opciones.nth(k).inner_text()
                txt_norm = normalizar_texto(txt_opcion)
                if txt_norm.startswith(item_norm) or item_norm in txt_norm:
                    hacer_clic_opcion(opciones.nth(k))
                    encontrado = True
                    time.sleep(0.4)
                    break

        if not encontrado and total > 0:
            hacer_clic_opcion(opciones.first)
            time.sleep(0.4)

    try:
        page.keyboard.press("Escape")
    except Exception:
        pass


def llenar_select(page, campo, valor, timeout_ms, app):
    valor_norm = normalizar_texto(valor)
    col_nombre = campo.get("columna", "")
    selector_config = campo.get("selector", "")

    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    if modal_activo.count() > 0:
        modal_top = modal_activo.last
    else:
        modal_top = page

    targets_posibles = [
        modal_top.locator(selector_config).last if selector_config else None,
        modal_top.locator(f".ant-form-item:has-text('{col_nombre}') .ant-select-selector").last,
        modal_top.locator(f".ant-form-item:has-text('{col_nombre}') input").last,
        modal_top.locator(f".ant-form-item:has-text('{col_nombre}')").last,
        page.locator(selector_config).last if selector_config else None,
        page.locator(selector_config).first if selector_config else None,
    ]

    dropdown_sel = ".ant-select-dropdown:not(.ant-select-dropdown-hidden)"

    menu_abierto = False
    for target in targets_posibles:
        if target is None:
            continue
        try:
            if target.count() > 0:
                target.scroll_into_view_if_needed(timeout=1000)
                target.click(force=True, timeout=2000)
                time.sleep(0.5)
                if page.locator(dropdown_sel).count() > 0:
                    menu_abierto = True
                    break
        except Exception:
            continue

    if not menu_abierto:
        try:
            modal_top.locator(f".ant-form-item:has-text('{col_nombre}')").last.click(force=True)
            time.sleep(0.6)
        except Exception:
            pass

    page.wait_for_selector(dropdown_sel, timeout=max(6000, timeout_ms))
    
    dropdown_activo = page.locator(dropdown_sel).last
    opciones = dropdown_activo.locator(".ant-select-item-option, div[role='option'], .ant-select-item-option-content")
    total_opciones = opciones.count()

    if total_opciones == 0:
        opciones = page.locator(dropdown_sel + " .ant-select-item-option")
        total_opciones = opciones.count()

    if total_opciones == 0:
        raise ValueError(f"No se desplegaron opciones en el menú para '{valor}'")

    app.log(f"      🔍 Se encontraron {total_opciones} opciones en el menú desplegable para '{valor}'.", "detail")

    for k in range(total_opciones):
        txt_opcion = opciones.nth(k).inner_text()
        txt_norm = normalizar_texto(txt_opcion)
        if txt_norm == valor_norm:
            app.log(f"      🎯 Seleccionada opción '{txt_opcion.strip()}' (para '{valor}')", "detail")
            hacer_clic_opcion(opciones.nth(k))
            time.sleep(0.6)
            return

    for k in range(total_opciones):
        txt_opcion = opciones.nth(k).inner_text()
        txt_norm = normalizar_texto(txt_opcion)
        if txt_norm.startswith(valor_norm) or valor_norm in txt_norm:
            app.log(f"      🎯 Seleccionada opción semejante '{txt_opcion.strip()}' (para '{valor}')", "detail")
            hacer_clic_opcion(opciones.nth(k))
            time.sleep(0.6)
            return

    try:
        opc_has = dropdown_activo.locator(f":has-text('{valor}')").last
        if opc_has.count() > 0:
            hacer_clic_opcion(opc_has)
            time.sleep(0.6)
            return
    except Exception:
        pass

    raise ValueError(f"La opción '{valor}' no se encuentra en la lista del menú activo")


def llenar_switch(page, campo, valor, app):
    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    if modal_activo.count() > 0:
        modal_top = modal_activo.last
        target = modal_top.locator(campo["selector"]).last
        if target.count() == 0:
            target = modal_top.locator("#isNanomaterial, button#isNanomaterial, button.ant-switch").last
    else:
        target = page.locator(campo["selector"]).last
        if target.count() == 0:
            target = page.locator("#isNanomaterial, button#isNanomaterial, .ant-modal-body button.ant-switch, button.ant-switch").last

    if target.count() == 0:
        app.log(f"   ⚠️ No se encontró el interruptor para '{campo['columna']}'", "warning")
        return

    deseado_si = str(valor).strip().lower() in ["sí", "si", "s", "true", "1", "yes"]

    try:
        aria_checked = target.get_attribute("aria-checked")
        clases = target.get_attribute("class") or ""
        esta_activo = aria_checked == "true" or "ant-switch-checked" in clases

        if deseado_si and not esta_activo:
            app.log(f"   🔘 Activando interruptor '{campo['columna']}' (Sí)", "detail")
            try:
                target.click(force=True)
            except Exception:
                target.evaluate("el => el.click()")
            time.sleep(0.4)
        elif not deseado_si and esta_activo:
            app.log(f"   🔘 Desactivando interruptor '{campo['columna']}' (No)", "detail")
            try:
                target.click(force=True)
            except Exception:
                target.evaluate("el => el.click()")
            time.sleep(0.3)
    except Exception as e:
        app.log(f"   ⚠️ No se pudo cambiar el interruptor '{campo['columna']}': {e}", "warning")


def _seleccionar_primera_opcion_listado_referencia(page, campo, app, timeout_ms=6000):
    # 1. Obtener el modal activo superior
    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    modal_top = modal_activo.last if modal_activo.count() > 0 else page

    # 2. Localizadores específicos del input de Listado de referencia
    candidatos = [
        modal_top.locator(".ant-form-item:has-text('Listado') input").last,
        modal_top.locator(".ant-form-item:has-text('referencia') input").last,
        modal_top.locator(".ant-form-item:has-text('Referencia') input").last,
        modal_top.locator("form > div > div:nth-child(4) input").last,
        modal_top.locator("#referenceList, input[id*='reference'], input[id*='Reference']").last,
        modal_top.locator(".ant-form-item:has-text('Listado') .ant-select-selector").last,
        modal_top.locator(".ant-form-item:has-text('referencia') .ant-select-selector").last,
        modal_top.locator("form > div > div:nth-child(4) .ant-select-selector").last,
        page.locator(".ant-modal-body .ant-form-item:has-text('Listado') input").last,
        page.locator(campo.get("selector", "")).last,
    ]

    target = None
    for cand in candidatos:
        try:
            if cand.count() > 0:
                target = cand
                break
        except Exception:
            continue

    if not target or target.count() == 0:
        app.log("   ⚠️ No se encontró el campo 'Listado de referencia' en el formulario activo.", "warning")
        raise Exception("El campo obligatorio 'Listado de referencia' no se encontró en el modal actual.")

    # Si el target no es input directo, buscar su input interno
    try:
        if target.evaluate("el => el.tagName.toLowerCase()") != "input":
            inp_hijo = target.locator("input").first
            if inp_hijo.count() > 0:
                target = inp_hijo
    except Exception:
        pass

    app.log("   ➔ Listado de referencia: (Vacío en Excel -> Disparando búsqueda y seleccionando primera opción)...", "detail")

    # Función auxiliar para hacer clic JavaScript en la primera opción visible del dropdown
    def intentar_clic_js():
        return page.evaluate("""() => {
            const dropdowns = Array.from(document.querySelectorAll('.ant-select-dropdown')).filter(d => {
                const style = window.getComputedStyle(d);
                return style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0';
            });
            if (dropdowns.length === 0) return null;
            const activeDd = dropdowns[dropdowns.length - 1];
            const opt = activeDd.querySelector('.ant-select-item-option');
            if (opt) {
                const text = (opt.innerText || opt.textContent || '').trim();
                opt.scrollIntoView();
                const content = opt.querySelector('.ant-select-item-option-content') || opt;
                content.dispatchEvent(new MouseEvent('mousedown', { bubbles: true, cancelable: true }));
                content.dispatchEvent(new MouseEvent('mouseup', { bubbles: true, cancelable: true }));
                content.click();
                return text;
            }
            return null;
        }""")

    # 3. Primer intento: Clic en el campo y flecha abajo
    try:
        target.scroll_into_view_if_needed(timeout=800)
        target.click(force=True, timeout=1000)
        time.sleep(0.3)
        page.keyboard.press("ArrowDown")
        time.sleep(0.3)
        res_js = intentar_clic_js()
        if res_js:
            app.log(f"      🎯 Seleccionada opción de referencia: '{res_js}'", "detail")
            time.sleep(0.6)
            return True
    except Exception:
        pass

    # 4. Segundo intento: Tipear letras clave para disparar la consulta al catálogo
    terminos = ["a", "d", "c", "i", "e", "CosIng", " "]
    for termino in terminos:
        try:
            target.click(force=True, timeout=1000)
            time.sleep(0.1)
            target.fill("")
            time.sleep(0.1)
            target.type(termino, delay=60)
            time.sleep(0.6)

            # Apenas escribe, chequear si el dropdown se abrió y hacer clic JS
            for _ in range(5):
                res_js = intentar_clic_js()
                if res_js:
                    app.log(f"      🎯 Seleccionada opción de referencia: '{res_js}'", "detail")
                    time.sleep(0.8)
                    return True
                time.sleep(0.3)

            # Si no hizo clic JS, probar inmediatamente ArrowDown y Enter sin volver a clickear el input
            page.keyboard.press("ArrowDown")
            time.sleep(0.2)
            page.keyboard.press("Enter")
            time.sleep(0.6)

            # Verificar si se seleccionó algo en el input
            val_input = target.evaluate("el => el.value") or ""
            if val_input and val_input != termino:
                app.log(f"      🎯 Opción confirmada por teclado: '{val_input}'", "detail")
                time.sleep(0.4)
                return True
        except Exception:
            continue

    # 5. Si ninguna búsqueda funcionó, verificar si hay opciones visibles y cliquearlas con Playwright
    try:
        opciones_visibles = page.locator(".ant-select-dropdown:visible .ant-select-item-option:visible, .ant-select-dropdown:visible .ant-select-item-option-content:visible")
        if opciones_visibles.count() > 0:
            txt_opc = opciones_visibles.first.inner_text().strip()
            app.log(f"      🎯 Clic Playwright en opción visible: '{txt_opc}'", "detail")
            opciones_visibles.first.click(force=True)
            time.sleep(0.8)
            return True
    except Exception:
        pass

    raise Exception("El campo 'Listado de referencia' es obligatorio y no se pudo seleccionar ninguna opción de la lista.")


def llenar_campo_composicion(page, campo, fila, app, cfg):
    col_nombre = campo["columna"]
    valor = fila.get(col_nombre, "")
    if not valor:
        for a in campo.get("alias", []):
            if a in fila and fila[a]:
                valor = fila[a]
                break

    es_listado_ref = "listado" in col_nombre.lower() or "referencia" in col_nombre.lower()
    timeout_ms = getattr(cfg, "TIMEOUT_SEGUNDOS", 15) * 1000

    if not valor or str(valor).strip() == "":
        if es_listado_ref:
            _seleccionar_primera_opcion_listado_referencia(page, campo, app, timeout_ms)
        return

    # Buscar target en el modal activo superior
    modal_activo = page.locator(".ant-modal:not([style*='display: none'])")
    if modal_activo.count() > 0:
        modal_top = modal_activo.last
        target = modal_top.locator(campo["selector"]).last
        if target.count() == 0:
            target = modal_top.locator(f".ant-form-item:has-text('{col_nombre}') input").last
        if target.count() == 0:
            target = modal_top.locator(f".ant-form-item:has-text('{col_nombre}') .ant-select-selector").last
        if target.count() == 0:
            target = page.locator(campo["selector"]).last
    else:
        modal_top = page
        target = page.locator(campo["selector"]).last

    if target.count() == 0:
        target = page.locator(campo["selector"]).first

    if target.count() == 0:
        app.log(f"   ℹ️ Campo '{campo['columna']}' no visible/encontrado, omitiendo...", "info")
        return

    # Esperar si el elemento está deshabilitado
    try:
        if target.is_disabled():
            app.log(f"   ⏳ Esperando a que el portal desbloquee '{campo['columna']}'...", "detail")
            page.wait_for_function("el => !el.disabled", arg=target.element_handle(), timeout=4000)
    except Exception:
        pass

    app.log(f"   ➔ {campo['columna']}: '{valor}'", "detail")

    try:
        if campo["tipo"] == "texto":
            try:
                tag_name = target.evaluate("el => el.tagName.toLowerCase()")
            except Exception:
                tag_name = "input"
            if tag_name not in ["input", "textarea"]:
                input_child = target.locator("input, textarea").first
                if input_child.count() > 0:
                    target = input_child
            target.fill(str(valor))
        elif campo["tipo"] == "select":
            llenar_select(page, campo, valor, timeout_ms, app)
        elif campo["tipo"] == "autocompletar":
            llenar_autocompletar_composicion(page, campo, valor, timeout_ms, app)
        elif campo["tipo"] == "multiselect":
            llenar_multiselect(page, campo, valor, timeout_ms, app)
        elif campo["tipo"] in ["switch", "toggle"]:
            llenar_switch(page, campo, valor, app)
    except Exception as e:
        raise Exception(f"Columna '{campo['columna']}' (Valor: '{valor}') -> {e}")


def llenar_campo(page, campo, fila, linea_excel, app, cfg, nombre_proceso=""):
    es_presentaciones_proc = "presentaci" in str(nombre_proceso).lower() or "informaci" in str(nombre_proceso).lower()
    if es_presentaciones_proc:
        llenar_campo_presentaciones(page, campo, fila, app, cfg)
    else:
        llenar_campo_composicion(page, campo, fila, app, cfg)


# --------------------------------------------------------------------------
#  Generador de Reportes de Error
# --------------------------------------------------------------------------
def guardar_reporte_errores(ruta_excel, nombre_proceso, lista_errores):
    carpeta_salida = os.path.dirname(ruta_excel) if ruta_excel else os.getcwd()
    ruta_txt = os.path.join(carpeta_salida, f"reporte_errores_{nombre_proceso.replace(' ', '_')}.txt")

    try:
        with open(ruta_txt, "w", encoding="utf-8") as f:
            f.write("==========================================================================" + "\n")
            f.write("         AUTOMATIZADOR INVIMA - REPORTE DE ERRORES DE REGISTRO            " + "\n")
            f.write(f" Proceso: {nombre_proceso}" + "\n")
            f.write(f" Fecha y Hora: {time.strftime('%Y-%m-%d %H:%M:%S')}" + "\n")
            f.write(f" Archivo Excel: {os.path.basename(ruta_excel)}" + "\n")
            f.write("==========================================================================" + "\n\n")

            for err in lista_errores:
                f.write(f"📍 LÍNEA EN EXCEL #{err['linea_excel']}\n")
                f.write(f"   • Registro N°: {err['numero_registro']}\n")
                f.write(f"   • Columna Afectada: {err['columna']}\n")
                f.write(f"   • Valor Buscado: {err['valor']}\n")
                f.write(f"   • Razón del Error: {err['detalle']}\n")
                f.write(f"   • Intentos Realizados: {err['intentos']}\n")
                f.write("-" * 65 + "\n\n")

        return ruta_txt
    except Exception as e:
        print(f"Error escribiendo reporte de errores: {e}")
        return None


# --------------------------------------------------------------------------
#  Manejador Especializado para Fórmula Marco (2 Niveles: Fórmula e Ingredientes)
# --------------------------------------------------------------------------
def ejecutar_proceso_formula_marco(page, proceso_cfg, filas, app, cfg, timeout_ms, ruta_excel):
    from itertools import groupby

    col_nombre_formula = proceso_cfg.get("COLUMNA_NOMBRE_FORMULA", "Nombre de la fórmula marco")
    selector_abrir_formula = proceso_cfg.get("BOTON_ABRIR_MODAL", "")
    selector_modal_formula = proceso_cfg.get("SELECTOR_MODAL_PRINCIPAL", ".ant-modal-wrap, .ant-modal-content")
    selector_campo_nombre = proceso_cfg.get("SELECTOR_CAMPO_NOMBRE_FORMULA", ".ant-modal-body form input, .ant-modal-body input")
    selector_anadir_ing = proceso_cfg.get("BOTON_ANADIR_INGREDIENTE", "button:has-text('Añadir ingrediente'), button:has-text('Agregar ingrediente')")
    selector_modal_ing = proceso_cfg.get("SELECTOR_MODAL_INGREDIENTE", ".ant-modal-wrap, .ant-modal-content")
    selector_guardar_ing = proceso_cfg.get("BOTON_GUARDAR_INGREDIENTE", ".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar')")
    selector_guardar_formula = proceso_cfg.get("BOTON_GUARDAR_FORMULA", ".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar')")
    campos_ingrediente = proceso_cfg.get("CAMPOS", [])

    # Detectar el nombre real de la columna en el Excel
    col_nombre_formula_real = col_nombre_formula
    if filas:
        for k in filas[0].keys():
            if k == "__linea_excel__": continue
            if normalizar_texto(k) == normalizar_texto(col_nombre_formula) or "formula" in normalizar_texto(k):
                col_nombre_formula_real = k
                break

    grupos_formulas = []
    for nombre_f, items in groupby(filas, key=lambda f: f.get(col_nombre_formula_real, "").strip()):
        if not nombre_f:
            nombre_f = "Fórmula Marco"
        grupos_formulas.append((nombre_f, list(items)))

    total_formulas = len(grupos_formulas)
    total_ingredientes = len(filas)
    ingredientes_procesados = 0
    app.actualizar_progreso(0, total_ingredientes)
    app.log(f"📋 Se detectaron {total_formulas} fórmulas marco para procesar ({total_ingredientes} ingredientes en total).", "info")

    exitosos = 0
    errores = 0
    lista_errores = []

    for idx_f, (nombre_formula, lista_ingredientes) in enumerate(grupos_formulas, start=1):
        if app.debe_detener:
            app.log("\n⏹️ Proceso detenido completamente por el usuario.", "warning")
            break

        while app.debe_pausar and not app.debe_detener:
            time.sleep(0.3)

        if app.debe_detener:
            break

        app.log(f"\n🏷️ [Fórmula {idx_f} de {total_formulas}] '{nombre_formula}' ({len(lista_ingredientes)} ingredientes)...", "header")

        # 1. Abrir modal principal de Fórmula Marco con hasta 3 intentos
        modal_abierto = False
        for intento_f in range(1, 4):
            try:
                if intento_f > 1:
                    app.log(f"   🔄 Intento {intento_f}/3 para abrir Fórmula '{nombre_formula}'...", "warning")
                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                    except Exception:
                        pass

                app.log(f"   🖱️ Abriendo ventana de Fórmula Marco...", "detail")
                btn_abrir_form = page.locator(selector_abrir_formula).first
                try:
                    btn_abrir_form.scroll_into_view_if_needed(timeout=1000)
                    btn_abrir_form.click(force=True, timeout=3000)
                except Exception:
                    try:
                        btn_abrir_form.evaluate("el => el.click()")
                    except Exception:
                        page.locator("button:has-text('Fórmula marco'), button:has-text('Formula marco'), button:has-text('Agregar fórmula'), button:has-text('Adicionar fórmula')").first.click(force=True, timeout=3000)
                
                page.wait_for_selector(selector_modal_formula, state="visible", timeout=timeout_ms)
                time.sleep(0.4)

                app.log(f"   ✍️ Asignando nombre: '{nombre_formula}'", "detail")
                campo_nombre = page.locator(selector_campo_nombre).first
                campo_nombre.click(force=True, timeout=3000)
                campo_nombre.fill(nombre_formula)
                time.sleep(0.3)
                modal_abierto = True
                break
            except Exception as e_f:
                app.log(f"   ⚠️ Error en intento {intento_f}/3 de abrir fórmula: {e_f}", "warning")
                if intento_f == 3:
                    errores += len(lista_ingredientes)
                    lista_errores.append({
                        "linea_excel": lista_ingredientes[0]["__linea_excel__"] if lista_ingredientes else 0,
                        "columna_afectada": "Cabecera Fórmula Marco",
                        "valor_excel": nombre_formula,
                        "motivo": str(e_f)
                    })

        if not modal_abierto:
            continue

        # 2. Llenar cada uno de los ingredientes en el submodal con 3 intentos individuales
        for num_ing, fila_ing in enumerate(lista_ingredientes, start=1):
            while app.debe_pausar and not app.debe_detener:
                time.sleep(0.3)

            ingredientes_procesados += 1
            app.actualizar_progreso(ingredientes_procesados, total_ingredientes)

            linea_ex = fila_ing["__linea_excel__"]
            ing_nombre = fila_ing.get("Ingrediente / Mezcla") or fila_ing.get("Ingrediente") or f"#{num_ing}"
            app.log(f"   🧪 Ingrediente {num_ing}/{len(lista_ingredientes)}: '{ing_nombre}' (Línea Excel #{linea_ex})...", "detail")

            exito_ing = False
            ultimo_err_ing = ""

            for intento_ing in range(1, 4):
                if intento_ing > 1:
                    app.log(f"      🔄 INTENTO {intento_ing}/3 para ingrediente '{ing_nombre}' (Línea #{linea_ex})...", "warning")
                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                    except Exception:
                        pass

                try:
                    app.log(f"      🖱️ Clic en 'Añadir ingrediente'...", "detail")
                    btn_anadir = page.locator(selector_anadir_ing).first
                    try:
                        btn_anadir.scroll_into_view_if_needed(timeout=1000)
                        btn_anadir.click(force=True, timeout=timeout_ms)
                    except Exception:
                        page.locator("button:has-text('Añadir'), button:has-text('Agregar'), button:has-text('Adicionar')").first.click(force=True)

                    time.sleep(0.5)

                    for campo in campos_ingrediente:
                        llenar_campo(page, campo, fila_ing, linea_ex, app, cfg, "Composición")

                    app.log(f"      💾 Guardando ingrediente...", "detail")
                    btn_guardar_ing = page.locator(selector_guardar_ing).last
                    btn_guardar_ing.click(force=True, timeout=timeout_ms)
                    time.sleep(0.8)

                    exito_ing = True
                    exitosos += 1
                    app.incrementar_exitos()
                    app.log(f"      ✅ Ingrediente '{ing_nombre}' registrado con éxito.", "success")
                    break
                except Exception as e_ing:
                    ultimo_err_ing = str(e_ing)
                    app.log(f"      ⚠️ Intento {intento_ing}/3 falló: {e_ing}", "warning")

            if not exito_ing:
                errores += 1
                app.incrementar_errores()
                app.log(f"      ❌ ERROR DEFINITIVO en Ingrediente '{ing_nombre}' tras 3 intentos.", "error")
                lista_errores.append({
                    "linea_excel": linea_ex,
                    "columna_afectada": "Registro Ingrediente Fórmula Marco",
                    "valor_excel": ing_nombre,
                    "motivo": ultimo_err_ing
                })

            # Si el usuario presionó Detener durante el llenado, aseguramos este ingrediente y guardamos la fórmula
            if app.debe_detener:
                app.log(f"   🛑 Detención solicitada: Ingrediente #{num_ing} asegurado. Procediendo a guardar Fórmula '{nombre_formula}'...", "warning")
                break

        # 3. Guardar SIEMPRE la fórmula marco para asegurar los datos en el portal
        try:
            app.log(f"   💾 Guardando Fórmula Marco '{nombre_formula}'...", "detail")
            btn_guardar_form = page.locator(selector_guardar_formula).first
            btn_guardar_form.click(force=True, timeout=timeout_ms)
            time.sleep(1.0)
            app.log(f"   ✨ Fórmula Marco '{nombre_formula}' guardada exitosamente.", "success")
        except Exception as e_form_save:
            app.log(f"   ⚠️ Error al guardar Fórmula Marco '{nombre_formula}': {e_form_save}", "warning")

        if app.debe_detener:
            app.log(f"\n⏹️ Operación '{nombre_formula}' guardada exitosamente. Proceso detenido de forma segura.", "warning")
            break

    if lista_errores:
        guardar_reporte_errores(ruta_excel, "Fórmula Marco", lista_errores)

    return exitosos, errores


# --------------------------------------------------------------------------
#  Manejador Especializado para Composición por Grupo (Grupo + Fórmula Marco + Ingredientes)
# --------------------------------------------------------------------------
def ejecutar_proceso_composicion_grupo(page, proceso_cfg, filas, app, cfg, timeout_ms, ruta_excel):
    from itertools import groupby

    selector_abrir = proceso_cfg.get("BOTON_ABRIR_MODAL", "")
    selector_modal_principal = proceso_cfg.get("SELECTOR_MODAL_PRINCIPAL", ".ant-modal-wrap, .ant-modal-content")
    boton_accion_previa = proceso_cfg.get("BOTON_ACCION_PREVIA", "")
    campos_cabecera = proceso_cfg.get("CAMPOS_CABECERA", [])
    selector_anadir_ing = proceso_cfg.get("BOTON_ANADIR_INGREDIENTE", "button:has-text('Añadir ingrediente'), button:has-text('Agregar ingrediente'), button:has-text('Adicionar ingrediente')")
    selector_modal_ing = proceso_cfg.get("SELECTOR_MODAL_INGREDIENTE", ".ant-modal-wrap, .ant-modal-content")
    selector_guardar_ing = proceso_cfg.get("BOTON_GUARDAR_INGREDIENTE", ".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar'), button:has-text('Aceptar')")
    selector_guardar_grupo = proceso_cfg.get("BOTON_GUARDAR_GRUPO", ".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar'), button:has-text('Aceptar')")
    campos_ingrediente = proceso_cfg.get("CAMPOS", [])

    # Determinar la clave de agrupación (Grupo / Nombre del grupo)
    col_grupo = "Grupo"
    if filas:
        for k in filas[0].keys():
            if k == "__linea_excel__": continue
            if "grupo" in normalizar_texto(k):
                col_grupo = k
                break

    # Agrupación inteligente para Composición por Grupo:
    # Soporta tanto si repiten el nombre del Grupo en cada fila como si dejan la celda en blanco hacia abajo (Forward Fill)
    from collections import OrderedDict
    grupos_map = OrderedDict()
    ultimo_grupo = None

    for fila in filas:
        val_g = (fila.get(col_grupo) or "").strip()
        if val_g:
            ultimo_grupo = val_g
        elif not ultimo_grupo:
            ultimo_grupo = "Grupo Principal"

        if ultimo_grupo not in grupos_map:
            grupos_map[ultimo_grupo] = []
        grupos_map[ultimo_grupo].append(fila)

    grupos_lista = list(grupos_map.items())

    total_grupos = len(grupos_lista)
    total_ingredientes = len(filas)
    ingredientes_procesados = 0
    app.actualizar_progreso(0, total_ingredientes)
    app.log(f"📋 Se detectaron {total_grupos} grupos de composición para procesar ({total_ingredientes} ingredientes en total).", "info")

    exitosos = 0
    errores = 0
    lista_errores = []

    for idx_g, (nombre_grupo, lista_ingredientes) in enumerate(grupos_lista, start=1):
        if app.debe_detener:
            app.log("\n⏹️ Proceso detenido de forma segura antes de iniciar el siguiente grupo.", "warning")
            break

        while app.debe_pausar and not app.debe_detener:
            time.sleep(0.3)

        if app.debe_detener:
            app.log("\n⏹️ Proceso detenido de forma segura antes de iniciar el siguiente grupo.", "warning")
            break

        app.log(f"\n👥 [Grupo {idx_g} de {total_grupos}] '{nombre_grupo}' ({len(lista_ingredientes)} ingredientes)...", "header")

        # 1. Abrir y configurar Grupo con hasta 3 intentos
        modal_grupo_abierto = False
        for intento_g in range(1, 4):
            try:
                if intento_g > 1:
                    app.log(f"   🔄 INTENTO {intento_g}/3 para inicializar Grupo '{nombre_grupo}'...", "warning")
                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                    except Exception:
                        pass

                app.log(f"   🖱️ Abriendo ventana 'Composición por Grupo'...", "detail")
                btn_abrir_grp = page.locator("button:has-text('Añadir Composición por Grupo'), button:has-text('Agregar Composición por Grupo'), button:has-text('Añadir composición por grupo'), button:has-text('Agregar composición por grupo'), button:has-text('Composición por Grupo'), button:has-text('Composicion por Grupo')").first
                try:
                    btn_abrir_grp.scroll_into_view_if_needed(timeout=1000)
                except Exception:
                    pass
                
                try:
                    btn_abrir_grp.evaluate("el => el.click()")
                except Exception:
                    btn_abrir_grp.click(force=True, timeout=3000)
                
                page.wait_for_selector(".ant-modal:not([style*='display: none']), .ant-modal-wrap:not([style*='display: none']), .ant-modal-confirm", state="visible", timeout=timeout_ms)
                time.sleep(0.5)

                # Clic en 'Usar Fórmula Marco'
                if boton_accion_previa:
                    app.log(f"   🖱️ Seleccionando opción 'Usar Fórmula Marco'...", "detail")
                    time.sleep(0.3)
                    btn_marco = page.locator("button:has-text('Usar Fórmula Marco'), button:has-text('Usar fórmula marco'), button:has-text('Usar Formula Marco'), button:has-text('Usar formula marco'), .ant-modal-confirm-body button:has-text('Marco'), .ant-modal-confirm-body button:has-text('Fórmula')").first
                    try:
                        btn_marco.evaluate("el => el.click()")
                    except Exception:
                        btn_marco.click(force=True, timeout=3000)

                    time.sleep(0.8)
                    page.wait_for_selector(".ant-modal-body form, .ant-modal:not(.ant-modal-confirm)", state="visible", timeout=timeout_ms)

                # Llenar campos de cabecera del grupo (Grupo, Fórmula Marco)
                primera_fila = lista_ingredientes[0]
                linea_cabecera = primera_fila["__linea_excel__"]
                for campo_c in campos_cabecera:
                    val = primera_fila.get(campo_c["columna"], "")
                    if not val:
                        for a in campo_c.get("alias", []):
                            if a in primera_fila and primera_fila[a]:
                                val = primera_fila[a]
                                break
                    if val:
                        app.log(f"   📝 Asignando {campo_c['columna']}: '{val}'", "detail")
                        llenar_campo(page, campo_c, primera_fila, linea_cabecera, app, cfg, "Composición")
                        time.sleep(0.4)

                modal_grupo_abierto = True
                break

            except Exception as e_grp:
                app.log(f"   ⚠️ Intento {intento_g}/3 falló al inicializar Grupo: {e_grp}", "warning")
                if intento_g == 3:
                    errores += len(lista_ingredientes)
                    app.log(f"   ❌ ERROR DEFINITIVO al inicializar Grupo '{nombre_grupo}' tras 3 intentos.", "error")
                    lista_errores.append({
                        "linea_excel": lista_ingredientes[0]["__linea_excel__"] if lista_ingredientes else 0,
                        "columna_afectada": "Cabecera Composición por Grupo",
                        "valor_excel": nombre_grupo,
                        "motivo": str(e_grp)
                    })

        if not modal_grupo_abierto:
            continue

        # 2. Añadir cada ingrediente del grupo con sistema de 3 reintentos individuales
        for num_ing, fila_ing in enumerate(lista_ingredientes, start=1):
            while app.debe_pausar and not app.debe_detener:
                time.sleep(0.3)

            ingredientes_procesados += 1
            app.actualizar_progreso(ingredientes_procesados, total_ingredientes)

            linea_ex = fila_ing["__linea_excel__"]
            ing_nombre = fila_ing.get("Ingrediente / Mezcla") or fila_ing.get("Ingrediente") or f"#{num_ing}"
            app.log(f"   🧪 Ingrediente {num_ing}/{len(lista_ingredientes)}: '{ing_nombre}' (Línea Excel #{linea_ex})...", "detail")

            exito_ing = False
            ultimo_err_ing = ""

            for intento_ing in range(1, 4):
                if intento_ing > 1:
                    app.log(f"      🔄 INTENTO {intento_ing}/3 para ingrediente '{ing_nombre}' (Línea #{linea_ex})...", "warning")
                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                    except Exception:
                        pass

                try:
                    app.log(f"      🖱️ Clic en 'Añadir ingrediente'...", "detail")
                    btn_anadir = page.locator(selector_anadir_ing).first
                    try:
                        btn_anadir.scroll_into_view_if_needed(timeout=1000)
                        btn_anadir.click(force=True, timeout=timeout_ms)
                    except Exception:
                        page.locator("button:has-text('Añadir ingrediente'), button:has-text('Agregar ingrediente'), button:has-text('Adicionar ingrediente'), button:has-text('Añadir'), button:has-text('Agregar')").first.click(force=True)

                    time.sleep(0.6)

                    # Llenar campos de ingrediente
                    for campo in campos_ingrediente:
                        llenar_campo(page, campo, fila_ing, linea_ex, app, cfg, "Composición")

                    # Guardar ingrediente (en el sub-modal de ingrediente)
                    app.log(f"      💾 Guardando ingrediente...", "detail")
                    btn_guardar_ing = page.locator(selector_guardar_ing).last
                    btn_guardar_ing.click(force=True, timeout=timeout_ms)
                    time.sleep(0.8)

                    exito_ing = True
                    exitosos += 1
                    app.incrementar_exitos()
                    app.log(f"      ✅ Ingrediente '{ing_nombre}' registrado con éxito.", "success")
                    break

                except Exception as e_ing:
                    ultimo_err_ing = str(e_ing)
                    app.log(f"      ⚠️ Intento {intento_ing}/3 falló: {e_ing}", "warning")

            if not exito_ing:
                errores += 1
                app.incrementar_errores()
                app.log(f"      ❌ ERROR DEFINITIVO en Ingrediente '{ing_nombre}' (Línea #{linea_ex}) tras 3 intentos.", "error")
                lista_errores.append({
                    "linea_excel": linea_ex,
                    "columna_afectada": "Registro Ingrediente",
                    "valor_excel": ing_nombre,
                    "motivo": ultimo_err_ing
                })

            # Si el usuario presionó Detener durante el llenado, aseguramos este ingrediente y guardamos el grupo completo
            if app.debe_detener:
                app.log(f"   🛑 Detención solicitada: Ingrediente #{num_ing} asegurado. Procediendo a guardar Grupo '{nombre_grupo}'...", "warning")
                break

        # 3. Guardar SIEMPRE el grupo completo para asegurar los datos en el portal
        try:
            app.log(f"   💾 Guardando Grupo '{nombre_grupo}'...", "detail")
            btn_guardar_grp = page.locator(selector_guardar_grupo).first
            btn_guardar_grp.click(force=True, timeout=timeout_ms)
            time.sleep(1.0)
            app.log(f"   ✨ Grupo '{nombre_grupo}' guardado exitosamente.", "success")
        except Exception as e_grp_save:
            app.log(f"   ⚠️ Error al guardar Grupo '{nombre_grupo}': {e_grp_save}", "warning")

        if app.debe_detener:
            app.log(f"\n⏹️ Operación '{nombre_grupo}' guardada exitosamente. Proceso detenido de forma segura.", "warning")
            break

    if lista_errores:
        guardar_reporte_errores(ruta_excel, "Composición por Grupo", lista_errores)

    return exitosos, errores


# --------------------------------------------------------------------------
#  Motor de ejecución principal
# --------------------------------------------------------------------------
def ejecutar(ruta_excel, nombre_proceso, app):
    cfg = cargar_config_dinamico()
    app.log(f"⚙️ Iniciando automatización para el proceso: '{nombre_proceso}'", "info")

    proceso_cfg = obtener_dict_proceso(cfg, nombre_proceso)

    selector_abrir = proceso_cfg.get("BOTON_ABRIR_MODAL", "").strip()
    if not selector_abrir or "PON_EL_SELECTOR" in selector_abrir:
        app.log(f"❌ ERROR: El proceso '{nombre_proceso}' no tiene configurado BOTON_ABRIR_MODAL en config.py", "error")
        app.finalizar_proceso(exito=False)
        return

    filas = leer_excel(ruta_excel, app, cfg, proceso_cfg)
    if filas is None:
        app.finalizar_proceso(exito=False)
        return

    total = len(filas)
    app.actualizar_progreso(0, total)
    timeout_ms = getattr(cfg, "TIMEOUT_SEGUNDOS", 15) * 1000

    try:
        with sync_playwright() as p:
            if not esta_puerto_abierto(PUERTO_CHROME):
                app.log("🌐 Chrome automatizado no detectado en puerto 9222. Iniciando automáticamente...", "warning")
                abrir_chrome_automatizado(app)
                time.sleep(3)

            app.log("📡 Conectando con Google Chrome (puerto 9222)...", "info")
            try:
                browser = p.chromium.connect_over_cdp(f"http://localhost:{PUERTO_CHROME}")
            except Exception:
                # Segundo intento tras relanzar
                if abrir_chrome_automatizado(app):
                    time.sleep(2)
                    try:
                        browser = p.chromium.connect_over_cdp(f"http://localhost:{PUERTO_CHROME}")
                    except Exception as e2:
                        app.log(f"❌ ERROR CRÍTICO: No se pudo conectar a Chrome tras relanzar: {e2}", "error")
                        app.finalizar_proceso(exito=False)
                        return
                else:
                    app.log("❌ ERROR CRÍTICO: No se pudo conectar a Chrome (Puerto 9222).", "error")
                    app.finalizar_proceso(exito=False)
                    return

            contexto = browser.contexts[0]
            page = contexto.pages[0] if contexto.pages else contexto.new_page()
            page.set_default_timeout(timeout_ms)

            app.log(f"🔗 Conectado exitosamente a Chrome.", "success")
            app.log(f"🌐 Pestaña activa: {page.url}", "info")
            app.log("=" * 60, "divider")

            # Manejadores especializados según TIPO_PROCESO
            tipo_proceso = proceso_cfg.get("TIPO_PROCESO")
            if tipo_proceso == "formula_marco":
                exitosos, errores = ejecutar_proceso_formula_marco(page, proceso_cfg, filas, app, cfg, timeout_ms, ruta_excel)
                app.log("=" * 60, "divider")
                app.log(f"🏁 PROCESO FINALIZADO: {exitosos} ingredientes registrados, {errores} fallidos.", "header")
                app.finalizar_proceso(exito=(errores == 0))
                return
            elif tipo_proceso == "composicion_grupo":
                exitosos, errores = ejecutar_proceso_composicion_grupo(page, proceso_cfg, filas, app, cfg, timeout_ms, ruta_excel)
                app.log("=" * 60, "divider")
                app.log(f"🏁 PROCESO FINALIZADO: {exitosos} ingredientes registrados, {errores} fallidos.", "header")
                app.finalizar_proceso(exito=(errores == 0))
                return

            exitosos = 0
            errores = 0
            lista_errores = []

            selector_modal = proceso_cfg.get("SELECTOR_MODAL", "").strip() or ".ant-modal-content, .ant-modal, .modal-dialog"
            selector_enviar = proceso_cfg.get("BOTON_ENVIAR", "").strip() or ".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar'), button[type='submit']"
            campos_proceso = proceso_cfg.get("CAMPOS", [])

            for numero, fila in enumerate(filas, start=1):
                linea_excel = fila["__linea_excel__"]

                if app.debe_detener:
                    app.log("\n⏹️ Proceso detenido completamente por el usuario.", "warning")
                    break

                while app.debe_pausar and not app.debe_detener:
                    time.sleep(0.3)

                if app.debe_detener:
                    app.log("\n⏹️ Proceso detenido completamente por el usuario.", "warning")
                    break

                app.log(f"📌 Procesando Registro {numero} de {total} (Línea Excel #{linea_excel})...", "header")
                app.actualizar_progreso(numero, total)

                # Sistema de reintentos por fila (3 intentos para todos los procesos)
                es_presentaciones_proc = "presentaci" in str(nombre_proceso).lower() or "informaci" in str(nombre_proceso).lower()
                max_intentos = 3
                exito_fila = False
                ultimo_detalle_error = ""

                for intento in range(1, max_intentos + 1):
                    if app.debe_detener:
                        break

                    if intento > 1:
                        app.log(f"   🔄 INTENTO {intento} de {max_intentos} para Línea Excel #{linea_excel}...", "warning")
                        try:
                            page.keyboard.press("Escape")
                            time.sleep(0.5)
                        except Exception:
                            pass

                    try:
                        app.log(f"   🖱️ Abrir ventana modal...", "detail")
                        page.locator(selector_abrir).first.click(timeout=timeout_ms)

                        app.log(f"   ⏳ Esperando ventana modal...", "detail")
                        page.wait_for_selector(selector_modal, state="visible", timeout=timeout_ms)
                        time.sleep(0.5)

                        # Si el proceso requiere una acción previa dentro del modal (ej. 'Usar Fórmula Marco')
                        boton_accion_previa = proceso_cfg.get("BOTON_ACCION_PREVIA", "").strip()
                        if boton_accion_previa:
                            app.log(f"   🖱️ Buscando opción interior (Usar Fórmula Marco)...", "detail")
                            time.sleep(0.4)
                            clic_exitoso = False

                            candidatos = [
                                "button:has-text('Usar Fórmula Marco')",
                                "button:has-text('Usar fórmula marco')",
                                "button:has-text('Usar Formula Marco')",
                                "button:has-text('Usar formula marco')",
                                "button:has-text('Fórmula Marco')",
                                "button:has-text('Formula Marco')",
                                "button:has-text('Marco')",
                                ".ant-modal-confirm-body button",
                                ".ant-modal-body button",
                                boton_accion_previa
                            ]

                            for cand in candidatos:
                                try:
                                    loc = page.locator(cand)
                                    cnt = loc.count()
                                    if cnt > 0:
                                        for idx_btn in range(cnt):
                                            btn_elem = loc.nth(idx_btn)
                                            txt_btn = btn_elem.inner_text().strip()
                                            if "marco" in txt_btn.lower() or "formula" in txt_btn.lower():
                                                app.log(f"   🎯 Clic en opción encontrada: '{txt_btn}'", "detail")
                                                btn_elem.scroll_into_view_if_needed(timeout=1000)
                                                btn_elem.click(force=True, timeout=3000)
                                                clic_exitoso = True
                                                break
                                        if clic_exitoso:
                                            break
                                        if cand == boton_accion_previa:
                                            loc.first.click(force=True, timeout=3000)
                                            clic_exitoso = True
                                            break
                                except Exception:
                                    continue

                            if not clic_exitoso:
                                app.log(f"   ⚠️ Intentando clic forzado en selector configurado...", "warning")
                                page.locator(boton_accion_previa).first.click(force=True, timeout=timeout_ms)

                            time.sleep(0.8)

                        # Llenar cada uno de los campos configurados para este proceso
                        for campo in campos_proceso:
                            if "PON_EL_SELECTOR" in campo["selector"]:
                                continue
                            llenar_campo(page, campo, fila, linea_excel, app, cfg, nombre_proceso)

                        app.log(f"   💾 Enviando y guardando formulario...", "detail")
                        btn_env = page.locator(selector_enviar).first
                        if btn_env.count() > 0 and btn_env.is_visible():
                            btn_env.click(force=True, timeout=timeout_ms)
                        else:
                            page.locator(".ant-modal-footer button.ant-btn-primary, button:has-text('Guardar'), button:has-text('Adicionar'), button[type='submit']").first.click(force=True)

                        app.log(f"   ⏳ Esperando cierre de la ventana...", "detail")
                        try:
                            page.wait_for_selector(selector_modal, state="hidden", timeout=4000)
                        except Exception:
                            errores_form = []
                            try:
                                loc_errs = page.locator(".ant-form-item-explain-error, .ant-form-item-explain, .ant-form-item-has-error")
                                for e_idx in range(loc_errs.count()):
                                    txt_err = loc_errs.nth(e_idx).inner_text().strip()
                                    if txt_err and txt_err not in errores_form:
                                        errores_form.append(txt_err)
                            except Exception:
                                pass

                            try:
                                page.keyboard.press("Escape")
                                time.sleep(0.5)
                            except Exception:
                                pass

                            if errores_form:
                                detalle_err = "Error en el portal de INVIMA: " + " | ".join(errores_form)
                            else:
                                detalle_err = "El formulario modal no se cerró tras hacer clic en Guardar (campos requeridos incompletos o inválidos)."
                            raise ValueError(detalle_err)

                        exito_fila = True
                        break  # Exit retry loop on success

                    except PWTimeout:
                        ultimo_detalle_error = "Tiempo de espera agotado al interactuar con el elemento o guardar modal."
                    except Exception as e:
                        ultimo_detalle_error = str(e)

                    try:
                        page.keyboard.press("Escape")
                        time.sleep(0.4)
                    except Exception:
                        pass

                # Evaluación de resultado de la fila
                if exito_fila:
                    exitosos += 1
                    app.incrementar_exitos()
                    app.log(f"   ✨ Línea Excel #{linea_excel} guardada con éxito.", "success")
                else:
                    errores += 1
                    app.incrementar_errores()

                    col_afectada = "Desconocida / Estructura Modal"
                    valor_afectado = "N/A"
                    if "Columna '" in ultimo_detalle_error:
                        try:
                            col_afectada = ultimo_detalle_error.split("Columna '")[1].split("'")[0]
                        except Exception:
                            pass
                    if "Valor en Excel: '" in ultimo_detalle_error:
                        try:
                            valor_afectado = ultimo_detalle_error.split("Valor en Excel: '")[1].split("'")[0]
                        except Exception:
                            pass

                    app.log(f"   ❌ LÍNEA EXCEL #{linea_excel} FALLÓ TRAS {max_intentos} INTENTOS", "error")
                    app.log(f"      📍 Columna afectada: {col_afectada}", "error")
                    app.log(f"      📄 Valor en Excel: '{valor_afectado}'", "error")
                    app.log(f"      ⚠️ Motivo: {ultimo_detalle_error}", "warning")

                    lista_errores.append({
                        "linea_excel": linea_excel,
                        "numero_registro": numero,
                        "columna": col_afectada,
                        "valor": valor_afectado,
                        "detalle": ultimo_detalle_error,
                        "intentos": max_intentos
                    })

                    app.log(f"   ⏩ Continuando automáticamente con el siguiente registro...", "info")

            app.log("=" * 60, "divider")
            if app.debe_detener:
                app.log(f"🏁 PROCESO DETENIDO POR EL USUARIO: {exitosos} exitosos, {errores} errores.", "warning")
            else:
                app.log(f"📊 RESUMEN FINAL [{nombre_proceso}]: {exitosos} exitosos, {errores} fallidos de {total} registros.", "header")

            # Reporte de errores
            if lista_errores:
                ruta_reporte = guardar_reporte_errores(ruta_excel, nombre_proceso, lista_errores)
                app.log(f"\n📄 REPORTE DE ERRORES GENERADO:", "error")
                app.log(f"   📍 Guardado en: {ruta_reporte}", "warning")
                app.log(f"   💡 Abre este archivo para ver la línea exacta de Excel a corregir.", "info")

    except Exception as e:
        app.log(f"❌ Error general en ejecución: {e}", "error")

    app.finalizar_proceso(exito=(exitosos > 0 and errores == 0))


# --------------------------------------------------------------------------
#  Interfaz Gráfica Profesional con Selector de Proceso (CustomTkinter)
# --------------------------------------------------------------------------
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Automatizador INVIMA PRO")
        self.geometry("860x780")
        self.minsize(820, 700)

        # Cargar icono de ventana
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        ruta_ico = os.path.join(base_dir, "app_icon.ico")
        if os.path.exists(ruta_ico):
            try:
                self.iconbitmap(ruta_ico)
            except Exception:
                pass

        # Variables de control
        self.cfg = cargar_config_dinamico()
        self.ruta_excel = None
        self.debe_pausar = False
        self.debe_detener = False
        self.en_ejecucion = False
        self.num_exitos = 0
        self.num_errores = 0
        self.licencia_info = None

        self._construir_interfaz()
        buscar_actualizaciones_github(self)
        self.after(300, self._verificar_licencia_inicial)

    def _verificar_licencia_inicial(self):
        datos_locales = leer_licencia_local()
        if datos_locales and "clave" in datos_locales:
            clave = datos_locales["clave"]
            valido, res = validar_licencia_firebase(clave)
            if valido:
                self.licencia_info = res
                self.log(f"🔑 LICENCIA ACTIVA: {res.get('empresa')} (Equipos: {res.get('equipos_usados')}/{res.get('max_equipos')})", "success")
                return

        # Si no hay licencia local o no es válida, pedir activación
        self.mostrar_modal_activacion_licencia()

    def mostrar_modal_activacion_licencia(self, mensaje_error_inicial=""):
        top = ctk.CTkToplevel(self)
        top.title("🔐 Activación de Licencia - Bot INVIMA")
        top.geometry("520x360")
        top.resizable(False, False)
        top.attributes("-topmost", True)
        top.grab_set()

        def _on_cerrar_sin_licencia():
            if not self.licencia_info:
                messagebox.showwarning(
                    "Licencia Requerida",
                    "Se requiere una licencia activa para utilizar el Automatizador INVIMA.\nLa aplicación se cerrará."
                )
                self.destroy()
                sys.exit(0)
            else:
                top.destroy()

        top.protocol("WM_DELETE_WINDOW", _on_cerrar_sin_licencia)

        lbl_title = ctk.CTkLabel(
            top,
            text="🔐 Activación de Licencia de Software",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="#F8FAFC"
        )
        lbl_title.pack(pady=(22, 6))

        lbl_sub = ctk.CTkLabel(
            top,
            text="Ingresa tu Clave de Licencia proporcionada por el proveedor\npara activar el Automatizador INVIMA en este equipo.",
            font=ctk.CTkFont(size=12),
            text_color="#94A3B8"
        )
        lbl_sub.pack(pady=(0, 16))

        entry_clave = ctk.CTkEntry(
            top,
            placeholder_text="XXXX-XXXX-XXXX-XXXX",
            width=380,
            height=42,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        entry_clave.pack(pady=8)

        lbl_status = ctk.CTkLabel(
            top,
            text=mensaje_error_inicial,
            font=ctk.CTkFont(size=12),
            text_color="#EF4444" if mensaje_error_inicial else "#94A3B8"
        )
        lbl_status.pack(pady=6)

        def _activar():
            clave_ingresada = entry_clave.get().strip().upper()
            if not clave_ingresada:
                lbl_status.configure(text="❌ Por favor ingresa tu clave de licencia.", text_color="#EF4444")
                return

            lbl_status.configure(text="⏳ Verificando licencia en línea...", text_color="#3B82F6")
            top.update()

            valido, res = validar_licencia_firebase(clave_ingresada)
            if valido:
                self.licencia_info = res
                top.destroy()
                self.log(f"🎉 ¡Licencia Activada Exitosamente! Empresa: {res.get('empresa')}", "success")
                messagebox.showinfo("Licencia Activada", f"¡Bienvenido {res.get('empresa')}!\n\nTu software ha quedado activado en este equipo.")
            else:
                lbl_status.configure(text=f"❌ {res}", text_color="#EF4444")

        btn_activar = ctk.CTkButton(
            top,
            text="🚀 Activar Licencia Ahora",
            font=ctk.CTkFont(weight="bold", size=14),
            fg_color="#3B82F6",
            hover_color="#2563EB",
            height=42,
            width=220,
            command=_activar
        )
        btn_activar.pack(pady=(12, 10))

    def mostrar_modal_actualizacion(self, data):
        version_n = data.get("version", "Nueva versión")
        novedades = data.get("novedades", "Mejoras generales y correcciones.")
        exe_url = data.get("exe_url", "")
        config_url = data.get("config_url", "")

        def _construir_dialogo():
            try:
                top = ctk.CTkToplevel(self)
                top.title("✨ Actualización Disponible")
                top.geometry("500x340")
                top.resizable(False, False)
                top.attributes("-topmost", True)
                top.grab_set()

                lbl = ctk.CTkLabel(
                    top,
                    text=f"🎉 ¡Nueva versión {version_n} disponible!",
                    font=ctk.CTkFont(family="Segoe UI", size=17, weight="bold"),
                    text_color="#10B981"
                )
                lbl.pack(pady=(18, 6))

                lbl_sub = ctk.CTkLabel(
                    top,
                    text="Una versión más reciente del Automatizador está lista para instalar.",
                    font=ctk.CTkFont(size=12),
                    text_color="#94A3B8"
                )
                lbl_sub.pack(pady=(0, 10))

                txt = ctk.CTkTextbox(top, width=450, height=140, fg_color="#0F172A", text_color="#E2E8F0")
                txt.pack(pady=4)
                txt.insert("0.0", f"Novedades:\n{novedades}")
                txt.configure(state="disabled")

                def _actualizar():
                    top.destroy()
                    threading.Thread(target=ejecutar_actualizacion_automatica, args=(exe_url, config_url, self), daemon=True).start()

                btn = ctk.CTkButton(
                    top,
                    text="🚀 Actualizar Ahora Automáticamente",
                    font=ctk.CTkFont(weight="bold"),
                    fg_color="#10B981",
                    hover_color="#059669",
                    height=38,
                    command=_actualizar
                )
                btn.pack(pady=14)
            except Exception as e:
                print(f"Error mostrando modal de actualización: {e}")

        self.after(500, _construir_dialogo)

    def mostrar_modal_manual(self):
        top = ctk.CTkToplevel(self)
        top.title("📖 Manual de Usuario y Cláusula Legal - Automatizador INVIMA")
        top.geometry("700x580")
        top.resizable(True, True)
        top.attributes("-topmost", True)

        lbl_title = ctk.CTkLabel(
            top,
            text="📖 Manual de Operación y Cláusula de Exención de Responsabilidad",
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            text_color="#F8FAFC"
        )
        lbl_title.pack(pady=(16, 6))

        txt_manual = ctk.CTkTextbox(
            top,
            width=660,
            height=440,
            fg_color="#0F172A",
            text_color="#E2E8F0",
            font=ctk.CTkFont(family="Consolas", size=12)
        )
        txt_manual.pack(pady=8, padx=16, fill="both", expand=True)

        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        ruta_txt = os.path.join(base_dir, "MANUAL_DE_USUARIO.txt")
        texto_contenido = ""

        if os.path.exists(ruta_txt):
            try:
                with open(ruta_txt, "r", encoding="utf-8") as f:
                    texto_contenido = f.read()
            except Exception:
                pass

        if not texto_contenido:
            texto_contenido = """========================================================================
             MANUAL DE USUARIO Y GUÍA DE OPERACIÓN
                   AUTOMATIZADOR INVIMA v1.1
========================================================================

1. REQUISITOS PREVIOS DEL SISTEMA
------------------------------------------------------------------------
- Sistema Operativo: Windows 10 o Windows 11.
- Navegador: Google Chrome instalado.
- Formato de Archivo: Planillas en formato Microsoft Excel (.xlsx).
- Licencia Activa: Clave de Licencia proporcionada por el proveedor.

2. PASOS PARA LA EJECUCIÓN DEL PROGRAMA
------------------------------------------------------------------------
Paso 1: Abre la aplicación "Automatizador INVIMA.exe".
Paso 2: Si es la primera vez, ingresa tu Clave de Licencia y presiona "Activar Licencia".
Paso 3: Haz clic en el botón "🌐 Abrir Chrome Bot" situado en la barra superior.
Paso 4: En la ventana de Chrome que se abre, ingresa al portal de INVIMA con tus credenciales y navega exactamente hasta el formulario del trámite a diligenciar.
Paso 5: Selecciona el Proceso a automatizar en la aplicación (ej: "Información General (Presentaciones)" o "Composición").
Paso 6: Haz clic en "📂 Cargar Excel (.xlsx)" y selecciona tu archivo de datos.
Paso 7: Haz clic en "🚀 Comenzar Automatización".

3. ESTRUCTURA Y REGLAS DE LAS HOJAS DE EXCEL
------------------------------------------------------------------------
⚠️ REGLA DE ORO 1: Los nombres de las Hojas (pestañas de Excel) deben ser EXACTOS.
⚠️ REGLA DE ORO 2: Los encabezados de las columnas en la Fila 1 deben coincidir al 100% con los requeridos.
⚠️ REGLA DE ORO 3: Los valores y textos ingresados en las celdas (ej: tipos de envase, materiales, unidades de medida, etc.) deben coincidir AL 100% con las opciones desplegables del portal INVIMA, incluyendo tildes, mayúsculas y espacios. Si el portal tiene "Cojín", en Excel DEBE decir "Cojín" (no "cojin" ni "Cojin").

--- PROCESO 1: Información General (Presentaciones) ---
Nombre exacto de la Hoja en Excel: Presentaciones comerciales
Encabezados requeridos en la Fila 1:
  • Columna: Contenido Neto
  • Columna: Unidad de Medida
  • Columna: Tipo de Envase Primario
  • Columna: Material del Envase Primario
  • Columna: Tipo de Envase Secundario
  • Columna: Material del Envase Secundario
  • Columna: Observaciones

--- PROCESO 2: Composición ---
Nombre exacto de la Hoja en Excel: Composición
Encabezados requeridos en la Fila 1:
  • Columna: Tipo
  • Columna: Ingrediente / Mezcla
  • Columna: Función
  • Columna: Listado de referencia
  • Columna: Cantidad
  • Columna: Unidad de medida
  • Columna: ¿Es nanomaterial?
  • Columna: Tamaño de partícula (nm)

4. MANEJO DE ERRORES Y REPORTES
------------------------------------------------------------------------
- Si una fila falla en el Excel, el bot realizará hasta 3 reintentos automáticos por fila.
- Si tras 3 intentos no se logra registrar la fila, el bot guardará un reporte detallado en un archivo "reporte_errores_YYYY-MM-DD.txt" con el número de fila exacto y la causa, y continuará con las siguientes filas sin detener la ejecución.

5. CLÁUSULA DE EXENCIÓN DE RESPONSABILIDAD LEGAL (DISCLAIMER)
------------------------------------------------------------------------
• El software Automatizador INVIMA es una herramienta de asistencia y automatización robótica de tareas (RPA).
• EL DESARROLLADOR / PROVEEDOR NO SE HACE RESPONSABLE por mal uso de la herramienta, datos mal digitados por el usuario, errores en la plantilla Excel, nombres de encabezados incorrectos, inconsistencias en los registros ante la entidad INVIMA o multas/sanciones derivadas de información errónea ingresada por el usuario.
• Es responsabilidad exclusiva del usuario verificar y validar la exactitud de los datos ingresados en el archivo Excel y en la plataforma de INVIMA antes y después de ejecutar el proceso de automatización.
========================================================================
"""

        txt_manual.insert("0.0", texto_contenido)
        txt_manual.configure(state="disabled")

        def _abrir_txt():
            if os.path.exists(ruta_txt):
                os.startfile(ruta_txt)
            else:
                with open(ruta_txt, "w", encoding="utf-8") as f:
                    f.write(texto_contenido)
                os.startfile(ruta_txt)

        btn_abrir = ctk.CTkButton(
            top,
            text="📄 Abrir / Guardar Archivo TXT",
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#3B82F6",
            hover_color="#2563EB",
            command=_abrir_txt
        )
        btn_abrir.pack(pady=10)

    def _obtener_lista_procesos(self):
        if hasattr(self.cfg, "PROCESOS") and isinstance(self.cfg.PROCESOS, dict):
            return list(self.cfg.PROCESOS.keys())
        return ["Información General (Presentaciones)"]

    def _obtener_icono_proceso(self, nombre_proceso):
        nombre_lower = str(nombre_proceso).lower()
        if "presentaci" in nombre_lower:
            return "📦"
        elif "grupo" in nombre_lower and "composici" in nombre_lower:
            return "🔗"
        elif "grupo" in nombre_lower:
            return "👥"
        elif "marco" in nombre_lower:
            return "🧬"
        elif "ingrediente" in nombre_lower:
            return "🧪"
        elif "organol" in nombre_lower:
            return "👃"
        return "📋"

    def _obtener_hoja_para(self, nombre_proceso):
        proceso_cfg = obtener_dict_proceso(self.cfg, nombre_proceso)
        return proceso_cfg.get("NOMBRE_HOJA", "Matriz")

    def _obtener_hoja_actual(self):
        nombre_proceso = getattr(self, 'proceso_seleccionado', self._obtener_lista_procesos()[0])
        proceso_cfg = obtener_dict_proceso(self.cfg, nombre_proceso)
        return proceso_cfg.get("NOMBRE_HOJA", "Matriz Presentaciones")

    def _obtener_info_campos(self):
        nombre_proceso = getattr(self, 'proceso_seleccionado', self._obtener_lista_procesos()[0])
        proceso_cfg = obtener_dict_proceso(self.cfg, nombre_proceso)
        campos = proceso_cfg.get("CAMPOS", [])
        return f"{len(campos)} campos configurados"

    def seleccionar_proceso(self, nuevo_proceso):
        if self.en_ejecucion:
            return
        self.proceso_seleccionado = nuevo_proceso
        if hasattr(self, 'dropdown_menu'):
            self.dropdown_menu.actualizar_seleccion(nuevo_proceso)
        self.on_proceso_changed(nuevo_proceso)

    def _construir_interfaz(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)

        lista_procesos = self._obtener_lista_procesos()
        self.proceso_seleccionado = lista_procesos[0]

        # Componente de Dropdown Flotante Moderno (No desplaza la ventana, tiene barra lateral)
        class ModernFloatingDropdown(ctk.CTkFrame):
            def __init__(self, master, app, procesos):
                super().__init__(master, fg_color="transparent")
                self.app = app
                self.procesos = procesos
                self.popup = None

                # Botón Principal Azul Llamativo
                self.btn_principal = ctk.CTkButton(
                    self,
                    text=f"{self.app._obtener_icono_proceso(self.app.proceso_seleccionado)}  {self.app.proceso_seleccionado}   ▼",
                    font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                    fg_color="#2563EB",
                    hover_color="#1D4ED8",
                    border_color="#3B82F6",
                    border_width=1,
                    text_color="#FFFFFF",
                    corner_radius=10,
                    height=44,
                    anchor="w",
                    command=self.toggle
                )
                self.btn_principal.pack(fill="x", padx=16, pady=(0, 4))

            def toggle(self):
                if self.app.en_ejecucion:
                    return
                if self.popup and self.popup.winfo_exists():
                    self.cerrar()
                else:
                    self.abrir()

            def abrir(self):
                if self.popup and self.popup.winfo_exists():
                    try:
                        self.popup.destroy()
                    except Exception:
                        pass

                self.btn_principal.update_idletasks()
                x = self.btn_principal.winfo_rootx()
                y = self.btn_principal.winfo_rooty() + self.btn_principal.winfo_height() + 4
                ancho = self.btn_principal.winfo_width()
                alto_total = min(len(self.procesos) * 52 + 16, 260)

                self.popup = ctk.CTkToplevel(self.app)
                self.popup.overrideredirect(True)
                self.popup.attributes("-topmost", True)
                self.popup.geometry(f"{ancho}x{alto_total}+{x}+{y}")
                self.popup.configure(fg_color="#090D16")

                frame_borde = ctk.CTkFrame(
                    self.popup,
                    fg_color="#090D16",
                    corner_radius=12,
                    border_width=2,
                    border_color="#3B82F6"
                )
                frame_borde.pack(fill="both", expand=True)

                scroll_frame = ctk.CTkScrollableFrame(
                    frame_borde,
                    fg_color="transparent",
                    corner_radius=10,
                    scrollbar_button_color="#334155",
                    scrollbar_button_hover_color="#3B82F6"
                )
                scroll_frame.pack(fill="both", expand=True, padx=4, pady=4)

                for proc in self.procesos:
                    icono = self.app._obtener_icono_proceso(proc)
                    hoja = self.app._obtener_hoja_para(proc)
                    es_activo = (proc == self.app.proceso_seleccionado)

                    btn_item = ctk.CTkButton(
                        scroll_frame,
                        text=f"  {icono}  {proc}{'  ✓' if es_activo else ''}\n     📄 Hoja en Excel: {hoja}",
                        font=ctk.CTkFont(family="Segoe UI", size=12),
                        fg_color="#1D4ED8" if es_activo else "#090D16",
                        hover_color="#1E293B",
                        border_color="#38BDF8" if es_activo else "#1E293B",
                        border_width=1 if es_activo else 0,
                        text_color="#FFFFFF" if es_activo else "#CBD5E1",
                        corner_radius=8,
                        height=42,
                        anchor="w",
                        command=lambda p=proc: self.seleccionar(p)
                    )
                    btn_item.pack(fill="x", padx=4, pady=3)

                self.btn_principal.configure(
                    text=f"{self.app._obtener_icono_proceso(self.app.proceso_seleccionado)}  {self.app.proceso_seleccionado}   ▲"
                )

                self.app.bind("<Button-1>", self._on_app_click, add="+")

            def _on_app_click(self, event):
                if not self.popup or not self.popup.winfo_exists():
                    return
                x_click = event.x_root
                y_click = event.y_root
                try:
                    px = self.popup.winfo_rootx()
                    py = self.popup.winfo_rooty()
                    pw = self.popup.winfo_width()
                    ph = self.popup.winfo_height()
                    bx = self.btn_principal.winfo_rootx()
                    by = self.btn_principal.winfo_rooty()
                    bw = self.btn_principal.winfo_width()
                    bh = self.btn_principal.winfo_height()

                    if not (px <= x_click <= px + pw and py <= y_click <= py + ph) and \
                       not (bx <= x_click <= bx + bw and by <= y_click <= by + bh):
                        self.cerrar()
                except Exception:
                    pass

            def cerrar(self):
                if self.popup and self.popup.winfo_exists():
                    try:
                        self.popup.destroy()
                    except Exception:
                        pass
                    self.popup = None
                self.btn_principal.configure(
                    text=f"{self.app._obtener_icono_proceso(self.app.proceso_seleccionado)}  {self.app.proceso_seleccionado}   ▼"
                )

            def seleccionar(self, proc):
                self.app.seleccionar_proceso(proc)
                self.cerrar()

            def actualizar_seleccion(self, proc):
                self.btn_principal.configure(
                    text=f"{self.app._obtener_icono_proceso(proc)}  {proc}   ▼"
                )

            def configure_state(self, state):
                self.btn_principal.configure(state=state)
                if state == "disabled":
                    self.cerrar()

        # Proxy para compatibilidad con código existente que use self.opt_proceso
        class SelectorProcesoProxy:
            def __init__(self, app):
                self.app = app
            def get(self):
                return self.app.proceso_seleccionado
            def set(self, valor):
                self.app.seleccionar_proceso(valor)
            def configure(self, **kwargs):
                state = kwargs.get("state")
                if state and hasattr(self.app, "dropdown_menu"):
                    self.app.dropdown_menu.configure_state(state)

        self.opt_proceso = SelectorProcesoProxy(self)

        # 1. HEADER / BARRA DE TÍTULO PRINCIPAL (PREMIUM GLASSMORPHIC)
        frame_header = ctk.CTkFrame(self, fg_color="#0F172A", corner_radius=16, border_width=1, border_color="#1E293B")
        frame_header.grid(row=0, column=0, padx=18, pady=(16, 6), sticky="ew")
        frame_header.grid_columnconfigure(0, weight=1)

        frame_header_text = ctk.CTkFrame(frame_header, fg_color="transparent")
        frame_header_text.grid(row=0, column=0, padx=16, pady=10, sticky="w")

        lbl_title = ctk.CTkLabel(
            frame_header_text,
            text="⚡ AUTOMATIZADOR INVIMA PRO",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="#F8FAFC"
        )
        lbl_title.pack(anchor="w")

        lbl_subtitle = ctk.CTkLabel(
            frame_header_text,
            text="Motor Inteligente Multi-Proceso • Masterdent Soft",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#94A3B8"
        )
        lbl_subtitle.pack(anchor="w")

        frame_header_actions = ctk.CTkFrame(frame_header, fg_color="transparent")
        frame_header_actions.grid(row=0, column=1, padx=16, pady=10, sticky="e")

        btn_manual = ctk.CTkButton(
            frame_header_actions,
            text="📖 Manual",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            fg_color="#4F46E5",
            hover_color="#4338CA",
            height=32,
            width=90,
            corner_radius=8,
            command=self.mostrar_modal_manual
        )
        btn_manual.pack(side="left", padx=4)

        btn_chrome = ctk.CTkButton(
            frame_header_actions,
            text="🌐 Chrome Bot",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            fg_color="#0284C7",
            hover_color="#0369A1",
            height=32,
            width=110,
            corner_radius=8,
            command=lambda: abrir_chrome_automatizado(self)
        )
        btn_chrome.pack(side="left", padx=4)

        self.badge_estado = ctk.CTkLabel(
            frame_header_actions,
            text="● EN ESPERA",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color="#94A3B8",
            fg_color="#1E293B",
            corner_radius=8,
            padx=10,
            pady=5
        )
        self.badge_estado.pack(side="left", padx=(4, 0))

        # 2. SELECTOR DE FORMATO / PROCESO (MODERN DROPDOWN MENU)
        frame_proceso = ctk.CTkFrame(self, fg_color="#0F172A", corner_radius=16, border_width=1, border_color="#1E293B")
        frame_proceso.grid(row=1, column=0, padx=18, pady=4, sticky="ew")
        frame_proceso.grid_columnconfigure(0, weight=1)

        frame_proceso_header = ctk.CTkFrame(frame_proceso, fg_color="transparent")
        frame_proceso_header.pack(fill="x", padx=16, pady=(10, 4))

        lbl_sec_proceso = ctk.CTkLabel(
            frame_proceso_header,
            text="📋 FORMATO A REGISTRAR:",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color="#E2E8F0"
        )
        lbl_sec_proceso.pack(side="left")

        lbl_sec_hint = ctk.CTkLabel(
            frame_proceso_header,
            text="(Haz clic para desplegar opciones)",
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color="#64748B"
        )
        lbl_sec_hint.pack(side="left", padx=(6, 0))

        # Instancia del Dropdown Flotante Moderno
        self.dropdown_menu = ModernFloatingDropdown(frame_proceso, self, lista_procesos)
        self.dropdown_menu.pack(fill="x", pady=(0, 4))

        # Banner informativo de la hoja y campos
        frame_info_hoja = ctk.CTkFrame(frame_proceso, fg_color="#1E293B", corner_radius=8)
        frame_info_hoja.pack(fill="x", padx=16, pady=(0, 10))

        self.lbl_info_hoja = ctk.CTkLabel(
            frame_info_hoja,
            text=f"📌 Hoja en Excel: '{self._obtener_hoja_actual()}'",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color="#38BDF8",
            padx=10,
            pady=5
        )
        self.lbl_info_hoja.pack(side="left")

        self.lbl_info_campos = ctk.CTkLabel(
            frame_info_hoja,
            text=f"🧪 {self._obtener_info_campos()}",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#34D399",
            padx=10,
            pady=5
        )
        self.lbl_info_campos.pack(side="right")

        # 3. PANEL DE SELECCIÓN DE ARCHIVO EXCEL
        frame_top = ctk.CTkFrame(self, fg_color="#0F172A", corner_radius=16, border_width=1, border_color="#1E293B")
        frame_top.grid(row=2, column=0, padx=18, pady=4, sticky="ew")
        frame_top.grid_columnconfigure(0, weight=1)

        lbl_sec_archivo = ctk.CTkLabel(
            frame_top,
            text="📁 ARCHIVO EXCEL DE ENTRADA:",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color="#CBD5E1"
        )
        lbl_sec_archivo.grid(row=0, column=0, columnspan=2, padx=16, pady=(10, 2), sticky="w")

        self.entry_path = ctk.CTkEntry(
            frame_top,
            placeholder_text="Haz clic en 'Seleccionar Excel' para cargar tu matriz...",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            height=36,
            fg_color="#090D16",
            border_color="#334155",
            corner_radius=8
        )
        self.entry_path.grid(row=1, column=0, padx=(16, 8), pady=(0, 10), sticky="ew")

        self.btn_browse = ctk.CTkButton(
            frame_top,
            text="📂 Seleccionar Excel",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            fg_color="#3B82F6",
            hover_color="#2563EB",
            height=36,
            corner_radius=8,
            command=self.seleccionar_excel
        )
        self.btn_browse.grid(row=1, column=1, padx=(0, 16), pady=(0, 10))

        # 4. PANEL DE CONTROLES Y TELEMETRÍA EN TIEMPO REAL
        frame_controls = ctk.CTkFrame(self, fg_color="#0F172A", corner_radius=16, border_width=1, border_color="#1E293B")
        frame_controls.grid(row=3, column=0, padx=18, pady=4, sticky="ew")
        frame_controls.grid_columnconfigure((0, 1, 2), weight=1)

        self.btn_comenzar = ctk.CTkButton(
            frame_controls,
            text="🚀 Comenzar",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color="#10B981",
            hover_color="#059669",
            height=40,
            corner_radius=10,
            command=self.action_comenzar_reanudar
        )
        self.btn_comenzar.grid(row=0, column=0, padx=(16, 6), pady=(10, 6), sticky="ew")

        self.btn_pausar = ctk.CTkButton(
            frame_controls,
            text="⏸️ Pausar",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color="#F59E0B",
            hover_color="#D97706",
            height=40,
            corner_radius=10,
            state="disabled",
            command=self.action_pausar
        )
        self.btn_pausar.grid(row=0, column=1, padx=6, pady=(10, 6), sticky="ew")

        self.btn_detener = ctk.CTkButton(
            frame_controls,
            text="⏹️ Detener",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color="#EF4444",
            hover_color="#DC2626",
            height=40,
            corner_radius=10,
            state="disabled",
            command=self.action_detener
        )
        self.btn_detener.grid(row=0, column=2, padx=(6, 16), pady=(10, 6), sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(
            frame_controls,
            height=8,
            corner_radius=4,
            progress_color="#10B981",
            fg_color="#1E293B"
        )
        self.progress_bar.grid(row=1, column=0, columnspan=3, padx=16, pady=(0, 6), sticky="ew")
        self.progress_bar.set(0.0)

        frame_metrics = ctk.CTkFrame(frame_controls, fg_color="transparent")
        frame_metrics.grid(row=2, column=0, columnspan=3, padx=16, pady=(0, 8), sticky="ew")
        frame_metrics.grid_columnconfigure((0, 1, 2), weight=1)

        self.lbl_metric_filas = ctk.CTkLabel(
            frame_metrics, text="📊 Progreso: 0 / 0", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#94A3B8"
        )
        self.lbl_metric_filas.grid(row=0, column=0, sticky="w")

        self.lbl_metric_exitos = ctk.CTkLabel(
            frame_metrics, text="✅ Éxitos: 0", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#34D399"
        )
        self.lbl_metric_exitos.grid(row=0, column=1, sticky="n")

        self.lbl_metric_errores = ctk.CTkLabel(
            frame_metrics, text="❌ Errores: 0", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#F87171"
        )
        self.lbl_metric_errores.grid(row=0, column=2, sticky="e")

        # 5. TERMINAL / CONSOLA DE EVENTOS
        frame_log = ctk.CTkFrame(self, fg_color="#0F172A", corner_radius=16, border_width=1, border_color="#1E293B")
        frame_log.grid(row=4, column=0, padx=18, pady=(4, 14), sticky="nsew")
        frame_log.grid_columnconfigure(0, weight=1)
        frame_log.grid_rowconfigure(1, weight=1)

        lbl_log_title = ctk.CTkLabel(
            frame_log,
            text="💻 Consola de Registros y Eventos en Vivo",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color="#94A3B8"
        )
        lbl_log_title.grid(row=0, column=0, padx=16, pady=(8, 2), sticky="w")

        self.txt_log = ctk.CTkTextbox(
            frame_log,
            font=ctk.CTkFont(family="Consolas", size=11),
            fg_color="#030712",
            text_color="#38BDF8",
            corner_radius=10,
            border_width=1,
            border_color="#1E293B"
        )
        self.txt_log.grid(row=1, column=0, padx=12, pady=(0, 10), sticky="nsew")

    def on_proceso_changed(self, nuevo_proceso):
        self.cfg = cargar_config_dinamico()
        hoja_req = self._obtener_hoja_actual()
        campos_info = self._obtener_info_campos()
        self.lbl_info_hoja.configure(text=f"📌 Hoja en Excel: '{hoja_req}'")
        self.lbl_info_campos.configure(text=f"🧪 {campos_info}")
        self.log(f"🔄 Proceso cambiado a: '{nuevo_proceso}' (Hoja: '{hoja_req}')", "info")

        # Limpiar datos anteriores
        self.progress_bar.set(0.0)
        self.num_exitos = 0
        self.num_errores = 0
        self.lbl_metric_filas.configure(text="📊 Progreso: 0 / 0")
        self.lbl_metric_exitos.configure(text="✅ Éxitos: 0")
        self.lbl_metric_errores.configure(text="❌ Errores: 0")

    def set_estado_badge(self, texto, color_bg, color_txt="#FFFFFF"):
        self.badge_estado.configure(text=f"● {texto}", fg_color=color_bg, text_color=color_txt)

    def log(self, mensaje, tipo="normal"):
        timestamp = time.strftime("[%H:%M:%S] ")
        linea = f"{timestamp}{mensaje}\n"

        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", linea)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def seleccionar_excel(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel de Datos",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        if ruta:
            self.ruta_excel = ruta
            self.entry_path.delete(0, "end")
            self.entry_path.insert(0, ruta)
            self.log(f"📁 Archivo seleccionado: {os.path.basename(ruta)}", "info")

    def actualizar_progreso(self, actual, total):
        porcentaje = actual / total if total > 0 else 0
        def _update():
            try:
                self.progress_bar.set(porcentaje)
                self.lbl_metric_filas.configure(text=f"📊 Progreso: {actual} / {total}")
            except Exception:
                pass
        self.after(0, _update)

    def incrementar_exitos(self):
        self.num_exitos += 1
        def _update():
            try:
                self.lbl_metric_exitos.configure(text=f"✅ Éxitos: {self.num_exitos}")
            except Exception:
                pass
        self.after(0, _update)

    def incrementar_errores(self):
        self.num_errores += 1
        def _update():
            try:
                self.lbl_metric_errores.configure(text=f"❌ Errores: {self.num_errores}")
            except Exception:
                pass
        self.after(0, _update)

    def action_comenzar_reanudar(self):
        if not self.en_ejecucion:
            if not self.ruta_excel:
                self.log("⚠️ Por favor selecciona primero un archivo Excel válido.", "warning")
                return

            if not self.licencia_info or "clave" not in self.licencia_info:
                self.log("❌ ERROR CRÍTICO DE SEGURIDAD: No hay una licencia activa validada.", "error")
                messagebox.showerror("Licencia Requerida", "No puedes iniciar la automatización sin una licencia activa.")
                self.mostrar_modal_activacion_licencia("Debes activar una licencia para continuar.")
                return

            # Re-validación en tiempo real antes de iniciar
            clave = self.licencia_info["clave"]
            valido, res = validar_licencia_firebase(clave)
            if not valido:
                self.licencia_info = None
                self.log(f"❌ LICENCIA RECHAZADA POR EL SERVIDOR: {res}", "error")
                messagebox.showerror("Licencia Desactivada", f"La licencia ha sido inhabilitada o ha caducado:\n{res}")
                self.mostrar_modal_activacion_licencia(res)
                return

            nombre_proceso = self.opt_proceso.get()

            self.en_ejecucion = True
            self.debe_pausar = False
            self.debe_detener = False
            self.num_exitos = 0
            self.num_errores = 0
            self.lbl_metric_exitos.configure(text="✅ Éxitos: 0")
            self.lbl_metric_errores.configure(text="❌ Errores: 0")

            self.btn_browse.configure(state="disabled")
            self.opt_proceso.configure(state="disabled")
            self.btn_comenzar.configure(text="🚀 Ejecutando...", fg_color="#059669", state="disabled")
            self.btn_pausar.configure(state="normal", text="⏸️ Pausar", fg_color="#F59E0B")
            self.btn_detener.configure(state="normal")
            self.set_estado_badge("EJECUTANDO", "#059669")

            threading.Thread(target=ejecutar, args=(self.ruta_excel, nombre_proceso, self), daemon=True).start()

        elif self.debe_pausar:
            self.debe_pausar = False
            self.btn_comenzar.configure(text="🚀 Ejecutando...", fg_color="#059669", state="disabled")
            self.btn_pausar.configure(text="⏸️ Pausar", fg_color="#F59E0B")
            self.set_estado_badge("EJECUTANDO", "#059669")
            self.log("▶️ Proceso reanudado por el usuario.", "info")

    def action_pausar(self):
        if self.en_ejecucion and not self.debe_pausar:
            self.debe_pausar = True
            self.btn_pausar.configure(text="▶️ Reanudar", fg_color="#10B981")
            self.btn_comenzar.configure(text="▶️ Reanudar", fg_color="#10B981", state="normal")
            self.set_estado_badge("PAUSADO", "#D97706")
            self.log("⏸️ Proceso pausado. Haz clic en 'Reanudar' para continuar.", "warning")

    def action_detener(self):
        if self.en_ejecucion:
            self.debe_detener = True
            self.debe_pausar = False
            self.set_estado_badge("DETENIENDO...", "#DC2626")
            self.log("🛑 Cancelando proceso... Espere a finalizar la fila actual.", "warning")

    def finalizar_proceso(self, exito=True):
        self.en_ejecucion = False
        self.debe_pausar = False
        self.debe_detener = False

        self.btn_browse.configure(state="normal")
        self.opt_proceso.configure(state="normal")
        self.btn_comenzar.configure(text="🚀 Comenzar", fg_color="#10B981", state="normal")
        self.btn_pausar.configure(text="⏸️ Pausar", fg_color="#F59E0B", state="disabled")
        self.btn_detener.configure(state="disabled")

        if self.badge_estado.cget("text").endswith("DETENIENDO..."):
            self.set_estado_badge("DETENIDO", "#DC2626")
        elif exito:
            self.set_estado_badge("COMPLETADO", "#10B981")
        else:
            self.set_estado_badge("FINALIZADO CON ERRORES", "#EF4444")


if __name__ == "__main__":
    app = App()
    app.mainloop()
